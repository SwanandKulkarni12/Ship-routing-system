import asyncio
import heapq
import json
import logging
import os
import random
import time
import threading
import uuid
from datetime import datetime, timezone
from math import atan2, cos, radians, sin, sqrt
import math
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'), override=False)
import networkx as nx
import numpy as np
from scipy.interpolate import griddata
from build_subgraph import build_subgraph
from cost_calculation import calculate_weather_cost, compute_safety_risk
from marine_api import cmems_available, fetch_cmems_marine_grid, interpolate_cmems_at_point
from gfs_api import gfs_available, fetch_gfs_atmospheric_grid, build_hourly_weather_lookup_from_gfs
from graph_loader import build_spatial_index, find_k_nearest_water_nodes, find_nearest_water_node, load_navigation_graph
from vessel_polar import BASE_SPEED_KNOTS, calculate_eta_hours, calculate_fuel_and_co2, get_speed_factor_polar
from weather_api import OpenMeteoRateLimitError, fetch_weather_data_hourly, fetch_marine_data_hourly
import voyage_analyzer
weather_data = {}
marine_data = {}
R_NM = 3440.065
logging.basicConfig(level=os.getenv('LOG_LEVEL', 'INFO').upper(), format='%(asctime)s %(levelname)s %(name)s %(message)s')
logger = logging.getLogger(__name__)
GRAPH_PATH = os.path.join(os.path.dirname(__file__), 'grid_based_ship_routes.graphml')
GLOBAL_GRAPH = None
GLOBAL_TREE = None
GLOBAL_NODE_ARRAY = None
OBJECTIVE_PROFILES = {
    'balanced': {'alpha': 1.0, 'beta': 0.45, 'gamma': 1.0}, 
    'safety':   {'alpha': 0.7, 'beta': 1.5, 'gamma': 1.0}, 
    'distance': {'alpha': 1.0, 'beta': 0.0, 'gamma': 0.0}
}
MAX_ALLOWED_DETOUR_RATIO = 1.35
PARETO_WEIGHT_SETS = {'fastest': {'alpha': 1.0, 'beta': 0.0, 'gamma': 0.0}, 'safest': {'alpha': 0.0, 'beta': 1.0, 'gamma': 0.0}, 'efficient': {'alpha': 0.0, 'beta': 0.0, 'gamma': 1.0}}
FUEL_PRICE_PER_TONNE = float(os.getenv('FUEL_PRICE_PER_TONNE', '600.0'))
TIME_CHARTER_RATE_PER_HOUR = float(os.getenv('TIME_CHARTER_RATE_PER_HOUR', '2000.0'))
RISK_COST_PER_UNIT_RISK = float(os.getenv('RISK_COST_PER_UNIT_RISK', '300.0'))
FUEL_RATE_AT_REF_SPEED_TONNES_PER_NM = float(os.getenv('FUEL_RATE_AT_REF_SPEED_TONNES_PER_NM', '0.104'))
MAX_FUEL_SURGE_FACTOR = 2.5
MIN_FUEL_EFFICIENCY_FACTOR = 0.6
RISK_FREE_BAND = float(os.getenv('RISK_FREE_BAND', '0.20'))
RISK_COST_EXPONENT = float(os.getenv('RISK_COST_EXPONENT', '2.0'))
BALANCED_MAX_DISTANCE_INCREASE_PCT = float(os.getenv('BALANCED_MAX_DISTANCE_INCREASE_PCT', '3.0'))
BALANCED_MAX_ETA_INCREASE_HOURS = float(os.getenv('BALANCED_MAX_ETA_INCREASE_HOURS', '1.0'))
BALANCED_MAX_FUEL_INCREASE_PCT = float(os.getenv('BALANCED_MAX_FUEL_INCREASE_PCT', '2.0'))
BALANCED_MIN_RISK_GAIN_PTS_FOR_REGRESSION = float(os.getenv('BALANCED_MIN_RISK_GAIN_PTS_FOR_REGRESSION', '2.0'))
def _env_bool(name, default=False):
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {'1', 'true', 'yes', 'on'}
DEBUG_MODE = _env_bool('DEBUG_MODE', False)
RISK_STRICT_NO_GO_ALL_MODES = _env_bool('RISK_STRICT_NO_GO_ALL_MODES', True)
BALANCED_ACCEPTANCE_ENABLED = _env_bool('BALANCED_ACCEPTANCE_ENABLED', True)
OPEN_METEO_MAX_CALLS_PER_MIN = int(os.getenv('OPEN_METEO_MAX_CALLS_PER_MIN', '600'))
OPEN_METEO_MAX_CALLS_PER_HOUR = int(os.getenv('OPEN_METEO_MAX_CALLS_PER_HOUR', '5000'))
OPEN_METEO_MAX_CALLS_PER_DAY = int(os.getenv('OPEN_METEO_MAX_CALLS_PER_DAY', '10000'))
OPEN_METEO_BATCH_SIZE = int(os.getenv('OPEN_METEO_BATCH_SIZE', '150'))
OPEN_METEO_MAX_RETRIES = int(os.getenv('OPEN_METEO_MAX_RETRIES', '8'))
SAFETY_HARD_WEATHER_THRESHOLD = float(os.getenv('SAFETY_HARD_WEATHER_THRESHOLD', '0.65'))
SAFETY_PENALTY_MULTIPLIER = float(os.getenv('SAFETY_PENALTY_MULTIPLIER', '10.0'))
MAX_SAFE_WAVE_HEIGHT_M = float(os.getenv('MAX_SAFE_WAVE_HEIGHT_M', '6.0'))
MAX_SAFE_WIND_KMH = float(os.getenv('MAX_SAFE_WIND_KMH', '75.0'))
MAX_SAFE_CURRENT_KMH = float(os.getenv('MAX_SAFE_CURRENT_KMH', '8.0'))
SAFETY_BLOCK_ON_MISSING_WEATHER = _env_bool('SAFETY_BLOCK_ON_MISSING_WEATHER', False)
EDGE_WEATHER_SAMPLES = int(os.getenv('EDGE_WEATHER_SAMPLES', '3'))
CORRIDOR_RADIUS_KM = int(os.getenv('CORRIDOR_RADIUS_KM', '150'))
CORRIDOR_MIN_RADIUS_KM = int(os.getenv('CORRIDOR_MIN_RADIUS_KM', '400'))
CORRIDOR_MAX_RADIUS_KM = int(os.getenv('CORRIDOR_MAX_RADIUS_KM', '2000'))
WIND_FIELD_GRID_ROWS = int(os.getenv('WIND_FIELD_GRID_ROWS', '72'))
WIND_FIELD_GRID_COLS = int(os.getenv('WIND_FIELD_GRID_COLS', '72'))
WIND_GRID_RES_DEG = float(os.getenv('WIND_GRID_RES_DEG', '0.25'))
WIND_GRID_PADDING_DEG = float(os.getenv('WIND_GRID_PADDING_DEG', '0.5'))
WIND_GRID_INTERPOLATION_METHOD = os.getenv('WIND_GRID_INTERPOLATION_METHOD', 'linear').strip().lower()
WEATHER_MODEL_RUN_INTERVAL_HOURS = int(os.getenv('WEATHER_MODEL_RUN_INTERVAL_HOURS', '6'))
CORRIDOR_GRID_SPACING_DEG = float(os.getenv('CORRIDOR_GRID_SPACING_DEG', '0.25'))
CORRIDOR_GRID_PADDING_DEG = float(os.getenv('CORRIDOR_GRID_PADDING_DEG', '0.2'))
CORRIDOR_GRID_MAX_POINTS = int(os.getenv('CORRIDOR_GRID_MAX_POINTS', '1200'))
CORRIDOR_INTERP_NEIGHBORS = int(os.getenv('CORRIDOR_INTERP_NEIGHBORS', '6'))
ADAPTIVE_EDGE_SAMPLING_ENABLED = _env_bool('ADAPTIVE_EDGE_SAMPLING_ENABLED', True)
ROUTE_CACHE_MAX_ENTRIES = int(os.getenv('ROUTE_CACHE_MAX_ENTRIES', '64'))
WEATHER_CACHE_MAX_ENTRIES = int(os.getenv('WEATHER_CACHE_MAX_ENTRIES', '32'))
OPTIMIZATION_MAX_ITER = int(os.getenv('OPTIMIZATION_MAX_ITER', '3'))
OPTIMIZATION_CONVERGENCE_THRESHOLD = float(os.getenv('OPTIMIZATION_CONVERGENCE_THRESHOLD', '0.01'))
PARETO_ROUTE_ENABLED = _env_bool('PARETO_ROUTE_ENABLED', True)
SEVERITY_GRID_ROWS = int(os.getenv('SEVERITY_GRID_ROWS', '180'))
SEVERITY_GRID_COLS = int(os.getenv('SEVERITY_GRID_COLS', '180'))
SEVERITY_GRID_PADDING_DEG = float(os.getenv('SEVERITY_GRID_PADDING_DEG', '0.12'))
ENDPOINT_SNAP_TOP_K = int(os.getenv('ENDPOINT_SNAP_TOP_K', '3'))
ENDPOINT_PAIR_MAX_COMBINATIONS = int(os.getenv('ENDPOINT_PAIR_MAX_COMBINATIONS', '9'))
TURN_PENALTY_USD = float(os.getenv('TURN_PENALTY_USD', '0.0'))
TURN_PENALTY_DISTANCE_KM = float(os.getenv('TURN_PENALTY_DISTANCE_KM', '12.0'))
MAX_ROUTE_TURN_DEG = float(os.getenv('MAX_ROUTE_TURN_DEG', '165.0'))
CONSTRAINED_SMOOTH_POINTS_PER_EDGE = int(os.getenv('CONSTRAINED_SMOOTH_POINTS_PER_EDGE', '3'))
CONSTRAINED_SMOOTH_MAX_POINTS = int(os.getenv('CONSTRAINED_SMOOTH_MAX_POINTS', '300'))
ENDPOINT_CONNECTOR_MAX_STEP_KM = float(os.getenv('ENDPOINT_CONNECTOR_MAX_STEP_KM', '4.0'))
WEATHER_CUBE_CACHE = {}
ROUTE_RESPONSE_CACHE = {}
class AsyncRateLimiter:
    def __init__(self, calls_per_minute, calls_per_hour, calls_per_day=None):
        self.calls_per_minute = max(calls_per_minute, 1)
        self.calls_per_hour = max(calls_per_hour, 1)
        self.calls_per_day = max(int(calls_per_day), 1) if calls_per_day is not None else None
        self.calls = []
        self.lock = asyncio.Lock()
        self.extra_sleep = 0.0
    async def wait_for_slot(self, units=1):
        units = max(1, int(units))
        async with self.lock:
            now = time.time()
            minute_cutoff = now - 60
            hour_cutoff = now - 3600
            day_cutoff = now - 86400
            self.calls = [t for t in self.calls if t > day_cutoff]
            calls_last_min = [t for t in self.calls if t > minute_cutoff]
            calls_last_hour = [t for t in self.calls if t > hour_cutoff]
            sleep_seconds = 0.0
            if len(calls_last_min) + units > self.calls_per_minute:
                overflow = len(calls_last_min) + units - self.calls_per_minute
                idx = min(max(overflow - 1, 0), len(calls_last_min) - 1)
                sleep_seconds = max(sleep_seconds, 60 - (now - calls_last_min[idx]))
            if len(calls_last_hour) + units > self.calls_per_hour:
                overflow = len(calls_last_hour) + units - self.calls_per_hour
                idx = min(max(overflow - 1, 0), len(calls_last_hour) - 1)
                sleep_seconds = max(sleep_seconds, 3600 - (now - calls_last_hour[idx]))
            if self.calls_per_day is not None and len(self.calls) + units > self.calls_per_day:
                overflow = len(self.calls) + units - self.calls_per_day
                idx = min(max(overflow - 1, 0), len(self.calls) - 1)
                sleep_seconds = max(sleep_seconds, 86400 - (now - self.calls[idx]))
            if self.extra_sleep > 0:
                sleep_seconds = max(sleep_seconds, self.extra_sleep)
            if sleep_seconds > 0:
                logger.info('[rate-limiter] sleeping %.2fs before next batch', sleep_seconds)
                await asyncio.sleep(sleep_seconds)
            stamp = time.time()
            self.calls.extend([stamp] * units)
            self.extra_sleep = 0.0
    def on_rate_limit(self, retry_index, retry_after=None):
        jitter = random.uniform(0.0, 0.5)
        backoff = min(120.0, 2 ** retry_index + jitter)
        if retry_after is not None:
            backoff = max(backoff, float(retry_after))
        self.extra_sleep = min(120.0, max(self.extra_sleep, backoff))
        logger.warning('[rate-limiter] 429 detected retry=%s next_wait=%.1fs', retry_index + 1, self.extra_sleep)
RATE_LIMITER = AsyncRateLimiter(calls_per_minute=OPEN_METEO_MAX_CALLS_PER_MIN, calls_per_hour=OPEN_METEO_MAX_CALLS_PER_HOUR, calls_per_day=OPEN_METEO_MAX_CALLS_PER_DAY)
def haversine_distance(coord1, coord2):
    lon1, lat1 = coord1
    lon2, lat2 = coord2
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R_NM * c
def calculate_total_nautical_distance(path):
    total_distance = 0
    for i in range(len(path) - 1):
        dist = haversine_distance(path[i], path[i + 1])
        total_distance += dist
    return total_distance
def sample_path_locations_for_weather(path_lon_lat, max_points=40):
    if not path_lon_lat:
        return []
    stride = max(1, len(path_lon_lat) // max_points)
    sampled = []
    for i in range(0, len(path_lon_lat), stride):
        lon, lat = path_lon_lat[i]
        sampled.append(_quantize_location(lat, lon, precision=2))
    if path_lon_lat:
        lon, lat = path_lon_lat[-1]
        sampled.append(_quantize_location(lat, lon, precision=2))
    return list(dict.fromkeys(sampled))
async def choose_corridor_radius_km(a_star_path):
    sampled_locations = sample_path_locations_for_weather(a_star_path, max_points=45)
    if not sampled_locations:
        return (CORRIDOR_RADIUS_KM, {'max_wind_kmh': None, 'max_wave_m': None, 'reason': 'no_samples'})
    all_lats = [float(loc[0]) for loc in sampled_locations]
    all_lons = [float(loc[1]) for loc in sampled_locations]
    gfs_grid = await asyncio.to_thread(fetch_gfs_atmospheric_grid, min(all_lats), max(all_lats), min(all_lons), max(all_lons), 1)
    max_wind = 0.0
    if gfs_grid is not None:
        from gfs_api import interpolate_gfs_at_point
        for loc in sampled_locations:
            pt = interpolate_gfs_at_point(gfs_grid, float(loc[0]), float(loc[1]), hour=0)
            wind = pt.get('wind_speed_10m', 0.0)
            if wind:
                max_wind = max(max_wind, wind)
    if max_wind >= 120:
        radius = 2000
        reason = 'extreme_weather'
    elif max_wind >= 80:
        radius = 1000
        reason = 'severe_weather'
    elif max_wind >= 45:
        radius = 600
        reason = 'moderate_weather'
    else:
        radius = CORRIDOR_MIN_RADIUS_KM
        reason = 'normal_weather'
    radius = int(min(max(radius, CORRIDOR_MIN_RADIUS_KM), CORRIDOR_MAX_RADIUS_KM))
    return (radius, {'max_wind_kmh': round(max_wind, 2), 'max_wave_m': None, 'reason': reason})
def get_objective_profile(mode):
    if not mode:
        return (OBJECTIVE_PROFILES['balanced'], 'balanced')
    normalized = str(mode).strip().lower()
    aliases = {'lowest-risk': 'safety', 'risk': 'safety', 'fastest': 'distance', 'fuel-efficient': 'fuel consumption', 'fuel': 'fuel consumption'}
    normalized = aliases.get(normalized, normalized)
    if normalized in OBJECTIVE_PROFILES:
        return (OBJECTIVE_PROFILES[normalized], normalized)
    return (OBJECTIVE_PROFILES['balanced'], 'balanced')
def _calculate_geographic_bearing(pointA, pointB):
    lat1, lon1 = (np.radians(pointA[1]), np.radians(pointA[0]))
    lat2, lon2 = (np.radians(pointB[1]), np.radians(pointB[0]))
    dlon = lon2 - lon1
    x = np.sin(dlon) * np.cos(lat2)
    y = np.cos(lat1) * np.sin(lat2) - np.sin(lat1) * np.cos(lat2) * np.cos(dlon)
    initial_bearing = np.arctan2(x, y)
    return (np.degrees(initial_bearing) + 360) % 360
def _quantize_location(lat, lon, precision=3):
    return (round(float(lat), precision), round(float(lon), precision))
def _utc_model_run_key(interval_hours=WEATHER_MODEL_RUN_INTERVAL_HOURS):
    interval_hours = max(1, int(interval_hours))
    now_utc = datetime.now(timezone.utc)
    slot_hour = now_utc.hour // interval_hours * interval_hours
    slot_dt = now_utc.replace(hour=slot_hour, minute=0, second=0, microsecond=0)
    return slot_dt.strftime('%Y%m%d%H')
def _prune_cache_entries(cache_dict, max_entries):
    if len(cache_dict) <= max_entries:
        return
    ordered_keys = sorted(cache_dict.keys(), key=lambda key: cache_dict[key].get('created_at', 0.0))
    to_remove = len(cache_dict) - max_entries
    for key in ordered_keys[:to_remove]:
        cache_dict.pop(key, None)
def build_fixed_corridor_grid_points(nodes_lon_lat, max_points=CORRIDOR_GRID_MAX_POINTS):
    if not nodes_lon_lat:
        return []
    points = list(dict.fromkeys((_quantize_location(float(node[1]), float(node[0]), precision=2) for node in nodes_lon_lat)))
    return points
def _weather_cube_cache_key(subgraph, forecast_hours):
    nodes = list(subgraph.nodes())
    if not nodes:
        return None
    lons = [float(node[0]) for node in nodes]
    lats = [float(node[1]) for node in nodes]
    min_lon = round(min(lons), 1)
    max_lon = round(max(lons), 1)
    min_lat = round(min(lats), 1)
    max_lat = round(max(lats), 1)
    return (_utc_model_run_key(), min_lon, max_lon, min_lat, max_lat, round(CORRIDOR_GRID_SPACING_DEG, 3), int(CORRIDOR_GRID_MAX_POINTS), int(forecast_hours))
def estimate_node_arrival_hours(subgraph, end_node, total_distance_km):
    arrival_hours = {}
    if total_distance_km is None or not np.isfinite(total_distance_km) or total_distance_km <= 0:
        return arrival_hours
    try:
        reversed_graph = subgraph.reverse(copy=False)
        remaining_distances = nx.single_source_dijkstra_path_length(reversed_graph, end_node, weight='weight')
    except Exception:
        return arrival_hours
    for node, remaining_km in remaining_distances.items():
        if np.isfinite(remaining_km):
            dist_from_start_km = max(0.0, total_distance_km - remaining_km)
            arrival_hours[node] = int(dist_from_start_km / max(ROUTING_BASE_SPEED_KMH, 0.1))
    return arrival_hours
def _derive_current_from_hourly(grid_points, hourly_weather_lookup, hourly_marine_lookup):
    weather_lookup = {}
    marine_lookup = {}
    for loc in grid_points:
        weather_series = hourly_weather_lookup.get(loc) or []
        marine_series = hourly_marine_lookup.get(loc) or []
        weather_lookup[loc] = {'current': dict(weather_series[0]) if weather_series else {}}
        marine_lookup[loc] = {'current': dict(marine_series[0]) if marine_series else {}}
    weather_lookup, marine_lookup = impute_missing_weather_from_neighbors(grid_points, weather_lookup, marine_lookup)
    return (weather_lookup, marine_lookup)
def _build_weather_cube_arrays(grid_points, hourly_weather_lookup, hourly_marine_lookup, forecast_hours):
    n_points = len(grid_points)
    n_hours = max(1, int(forecast_hours))
    def to_matrix(lookup, field_name):
        matrix = np.full((n_points, n_hours), np.nan, dtype=np.float32)
        for idx, loc in enumerate(grid_points):
            series = lookup.get(loc)
            if not series:
                continue
            upper = min(len(series), n_hours)
            vals = [series[h].get(field_name) for h in range(upper)]
            arr = np.fromiter((float(v) if v is not None else np.nan for v in vals), dtype=np.float32, count=upper)
            matrix[idx, :upper] = arr
        return matrix
    weather_mats = {'wind_speed_10m': to_matrix(hourly_weather_lookup, 'wind_speed_10m'), 'wind_direction_10m': to_matrix(hourly_weather_lookup, 'wind_direction_10m'), 'precipitation': to_matrix(hourly_weather_lookup, 'precipitation'), 'visibility': to_matrix(hourly_weather_lookup, 'visibility')}
    marine_mats = {'wave_height': to_matrix(hourly_marine_lookup, 'wave_height'), 'wave_direction': to_matrix(hourly_marine_lookup, 'wave_direction'), 'ocean_current_velocity': to_matrix(hourly_marine_lookup, 'ocean_current_velocity'), 'ocean_current_direction': to_matrix(hourly_marine_lookup, 'ocean_current_direction')}
    return (weather_mats, marine_mats)
async def get_or_build_weather_context(subgraph, forecast_hours):
    cache_key = _weather_cube_cache_key(subgraph, forecast_hours)
    if cache_key and cache_key in WEATHER_CUBE_CACHE:
        cached = WEATHER_CUBE_CACHE[cache_key]
        cached['created_at'] = time.time()
        logger.info('[weather-cache] hit run=%s points=%s', cached['model_run_key'], len(cached['grid_points']))
        return cached
    grid_points = build_fixed_corridor_grid_points(list(subgraph.nodes()))
    if not grid_points:
        return {'model_run_key': _utc_model_run_key(), 'grid_points': [], 'hourly_weather_lookup': {}, 'hourly_marine_lookup': {}, 'current_weather_lookup': {}, 'current_marine_lookup': {}, 'weather_mats': {}, 'marine_mats': {}, 'lat_arr': np.array([], dtype=np.float32), 'lon_arr': np.array([], dtype=np.float32), 'interp_cache': {}, 'forecast_hours': int(forecast_hours), 'created_at': time.time(), 'severity_points': []}
    all_lats = [float(p[0]) for p in grid_points]
    all_lons = [float(p[1]) for p in grid_points]
    bbox_min_lat, bbox_max_lat = (min(all_lats), max(all_lats))
    bbox_min_lon, bbox_max_lon = (min(all_lons), max(all_lons))
    logger.info('[weather-cube] run=%s grid_points=%s forecast_hours=%s cmems=%s gfs=%s', _utc_model_run_key(), len(grid_points), forecast_hours, 'yes' if cmems_available() else 'no', 'yes' if gfs_available() else 'no')
    gfs_task = asyncio.to_thread(fetch_gfs_atmospheric_grid, bbox_min_lat, bbox_max_lat, bbox_min_lon, bbox_max_lon, int(forecast_hours))
    cmems_task = asyncio.to_thread(fetch_cmems_marine_grid, bbox_min_lat, bbox_max_lat, bbox_min_lon, bbox_max_lon, int(forecast_hours))
    gfs_grid, cmems_grid = await asyncio.gather(gfs_task, cmems_task)
    if gfs_grid is not None:
        logger.info('[gfs] building hourly weather lookup from GFS grid')
        hourly_weather_lookup = build_hourly_weather_lookup_from_gfs(grid_points, gfs_grid, int(forecast_hours))
    else:
        logger.warning('[gfs] fetch failed — falling back to Open-Meteo for wind/rain')
        hourly_weather_lookup, _, _ = await batch_fetch_weather_forecast(grid_points, forecast_hours=int(forecast_hours), batch_size=min(OPEN_METEO_BATCH_SIZE, 150))
    if cmems_grid is not None:
        logger.info('[cmems] building hourly marine lookup from CMEMS grid')
        fallback_marine = {'wave_height': 1.0, 'wave_direction': 180.0, 'ocean_current_velocity': 0.5, 'ocean_current_direction': 180.0}
        hourly_marine_lookup = {}
        for loc in grid_points:
            lat_pt, lon_pt = (float(loc[0]), float(loc[1]))
            cmems_pt = interpolate_cmems_at_point(cmems_grid, lat_pt, lon_pt)
            slot = {**fallback_marine, **cmems_pt} if cmems_pt else fallback_marine
            hourly_marine_lookup[loc] = [dict(slot)] * int(forecast_hours)
    else:
        logger.warning('[cmems] fetch failed — falling back to Open-Meteo for waves/currents')
        _, hourly_marine_lookup, _ = await batch_fetch_weather_forecast(grid_points, forecast_hours=int(forecast_hours), batch_size=min(OPEN_METEO_BATCH_SIZE, 150))
    severity_points = []
    for loc in grid_points:
        cw = hourly_weather_lookup.get(loc, [{}])[0]
        cm = hourly_marine_lookup.get(loc, [{}])[0]
        from cost_calculation import compute_safety_risk
        risk = round(compute_safety_risk(cw, cm) * 100.0, 2)
        severity_points.append({'coordinate': [loc[0], loc[1]], 'risk': risk})
    logger.info('[weather-cube] severity_pts=%s (from GFS+CMEMS hour-0)', len(severity_points))
    _tc = time.perf_counter()
    logger.info('[weather-cube] deriving current + building %sx%sh arrays in parallel', len(grid_points), forecast_hours)
    derive_task = asyncio.to_thread(_derive_current_from_hourly, grid_points, hourly_weather_lookup, hourly_marine_lookup)
    arrays_task = asyncio.to_thread(_build_weather_cube_arrays, grid_points, hourly_weather_lookup, hourly_marine_lookup, int(forecast_hours))
    (current_weather_lookup, current_marine_lookup), (weather_mats, marine_mats) = await asyncio.gather(derive_task, arrays_task)
    logger.info('[weather-cube] derive+arrays parallel: %.2fs', time.perf_counter() - _tc)
    context = {'model_run_key': _utc_model_run_key(), 'grid_points': grid_points, 'hourly_weather_lookup': hourly_weather_lookup, 'hourly_marine_lookup': hourly_marine_lookup, 'current_weather_lookup': current_weather_lookup, 'current_marine_lookup': current_marine_lookup, 'weather_mats': weather_mats, 'marine_mats': marine_mats, 'lat_arr': np.array([float(lat) for lat, _ in grid_points], dtype=np.float32), 'lon_arr': np.array([float(lon) for _, lon in grid_points], dtype=np.float32), 'interp_cache': {}, 'forecast_hours': int(forecast_hours), 'created_at': time.time(), 'severity_points': severity_points}
    if cache_key is not None:
        WEATHER_CUBE_CACHE[cache_key] = context
        _prune_cache_entries(WEATHER_CUBE_CACHE, WEATHER_CACHE_MAX_ENTRIES)
    return context
def interpolate_weather_from_context(weather_context, lat, lon, hour):
    if not weather_context or len(weather_context.get('grid_points', [])) == 0:
        return ({'current': {}}, {'current': {}})
    lat_q, lon_q = _quantize_location(lat, lon, precision=3)
    hour_idx = int(max(0, min(int(hour), weather_context['forecast_hours'] - 1)))
    cache_key = (lat_q, lon_q, hour_idx)
    interp_cache = weather_context.get('interp_cache', {})
    if cache_key in interp_cache:
        return interp_cache[cache_key]
    lat_arr = weather_context['lat_arr']
    lon_arr = weather_context['lon_arr']
    n_points = len(lat_arr)
    if n_points == 0:
        return ({'current': {}}, {'current': {}})
    dlat = lat_arr - float(lat_q)
    dlon = lon_arr - float(lon_q)
    dlon = (dlon + 180.0) % 360.0 - 180.0
    lon_scale = max(float(np.cos(np.radians(lat_q))), 0.2)
    dist2 = dlat * dlat + dlon * lon_scale * (dlon * lon_scale)
    k = min(max(1, CORRIDOR_INTERP_NEIGHBORS), n_points)
    nearest_idx = np.argpartition(dist2, k - 1)[:k]
    nearest_dist = np.sqrt(np.maximum(dist2[nearest_idx], 1e-12))
    weights = 1.0 / np.maximum(nearest_dist, 1e-06) ** 2
    def weighted_value(matrix):
        values = matrix[nearest_idx, hour_idx]
        valid = np.isfinite(values)
        if not np.any(valid):
            return None
        w = weights[valid]
        v = values[valid]
        return float(np.sum(w * v) / np.sum(w))
    weather_current = {}
    marine_current = {}
    for field_name, matrix in weather_context['weather_mats'].items():
        value = weighted_value(matrix)
        if value is not None:
            weather_current[field_name] = value
    for field_name, matrix in weather_context['marine_mats'].items():
        value = weighted_value(matrix)
        if value is not None:
            marine_current[field_name] = value
    weather_current.setdefault('wind_speed_10m', 10.0)
    weather_current.setdefault('wind_direction_10m', 180.0)
    weather_current.setdefault('precipitation', 0.0)
    weather_current.setdefault('visibility', 10000.0)
    marine_current.setdefault('wave_height', 1.0)
    marine_current.setdefault('wave_direction', 180.0)
    marine_current.setdefault('ocean_current_velocity', 0.5)
    marine_current.setdefault('ocean_current_direction', 180.0)
    payload = ({'current': weather_current}, {'current': marine_current})
    interp_cache[cache_key] = payload
    if len(interp_cache) > 120000:
        interp_cache.clear()
    return payload
_CW_ZERO_INVALID = {'wind_speed_10m', 'visibility'}
_CM_ZERO_INVALID = {'wave_height'}
def _field_needs_patch(field_name: str, value, zero_invalid_fields: set[str]) -> bool:
    if value is None:
        return True
    try:
        v = float(value)
    except (TypeError, ValueError):
        return True
    if not np.isfinite(v):
        return True
    return field_name in zero_invalid_fields and v == 0.0
def _missing_field_names(cw_fields: dict, cm_fields: dict) -> tuple[list[str], list[str]]:
    weather_missing = []
    marine_missing = []
    for f in ('wind_speed_10m', 'wind_direction_10m', 'precipitation', 'visibility'):
        if _field_needs_patch(f, cw_fields.get(f), _CW_ZERO_INVALID):
            weather_missing.append(f)
    for f in ('wave_height', 'wave_direction', 'ocean_current_velocity', 'ocean_current_direction'):
        if _field_needs_patch(f, cm_fields.get(f), _CM_ZERO_INVALID):
            marine_missing.append(f)
    return (weather_missing, marine_missing)
def _patch_with_fallback(cw_fields: dict, cm_fields: dict, fb_cw: dict, fb_cm: dict) -> tuple[dict, dict]:
    def _is_usable(val) -> bool:
        if val is None:
            return False
        try:
            return np.isfinite(float(val))
        except (TypeError, ValueError):
            return False
    cw = dict(cw_fields)
    cm = dict(cm_fields)
    for f in ('wind_speed_10m', 'wind_direction_10m', 'precipitation', 'visibility'):
        v = cw.get(f)
        if _field_needs_patch(f, v, _CW_ZERO_INVALID):
            fb_val = fb_cw.get(f)
            if _is_usable(fb_val):
                logger.info('patch field=%s  old=%s  new=%s', f, v, fb_val)
                cw[f] = fb_val
            else:
                logger.warning('patch field=%s  old=%s  fallback=%s unusable (NaN/None) — keeping original', f, v, fb_val)
    for f in ('wave_height', 'wave_direction', 'ocean_current_velocity', 'ocean_current_direction'):
        v = cm.get(f)
        if _field_needs_patch(f, v, _CM_ZERO_INVALID):
            fb_val = fb_cm.get(f)
            if _is_usable(fb_val):
                logger.info('patch field=%s  old=%s  new=%s', f, v, fb_val)
                cm[f] = fb_val
            else:
                logger.warning('patch field=%s  old=%s  fallback=%s unusable (NaN/None) — keeping original', f, v, fb_val)
    return (cw, cm)
def _batch_openmeteo_fallback_grouped(weather_locations: list[tuple[float, float]], marine_locations: list[tuple[float, float]], batch_size: int=OPEN_METEO_BATCH_SIZE) -> tuple[dict[tuple[float, float], dict], dict[tuple[float, float], dict]]:
    weather_locations = list(dict.fromkeys(weather_locations))
    marine_locations = list(dict.fromkeys(marine_locations))
    effective_batch = max(1, min(int(batch_size), 150))
    max_retries = max(1, min(OPEN_METEO_MAX_RETRIES, 5))
    weather_fallback_map = {}
    marine_fallback_map = {}
    if weather_locations:
        logger.info('build_weather_info: grouped Open-Meteo weather fallback points=%s batch_size=%s', len(weather_locations), effective_batch)
        for i in range(0, len(weather_locations), effective_batch):
            batch = weather_locations[i:i + effective_batch]
            lats = [loc[0] for loc in batch]
            lons = [loc[1] for loc in batch]
            weather_batch = None
            for retry_idx in range(max_retries):
                try:
                    weather_batch = fetch_weather_data_hourly(lats, lons, 1)
                    break
                except OpenMeteoRateLimitError as exc:
                    retry_after = exc.retry_after if exc.retry_after is not None else min(30.0, 2 ** retry_idx)
                    logger.warning('build_weather_info: weather fallback rate-limited retry=%s/%s sleep=%.1fs', retry_idx + 1, max_retries, float(retry_after))
                    time.sleep(float(retry_after))
                except Exception as exc:
                    logger.warning('build_weather_info: weather fallback batch failed error=%s', exc)
                    break
            for j, loc in enumerate(batch):
                slot = {}
                if weather_batch and j < len(weather_batch) and weather_batch[j]:
                    slot = dict(weather_batch[j][0])
                weather_fallback_map[loc] = slot
                logger.info('open-meteo weather fallback  lat=%.3f lon=%.3f  wind_speed=%s wind_dir=%s precip=%s visibility=%s', loc[0], loc[1], slot.get('wind_speed_10m'), slot.get('wind_direction_10m'), slot.get('precipitation'), slot.get('visibility'))
    if marine_locations:
        logger.info('build_weather_info: grouped Open-Meteo marine fallback points=%s batch_size=%s', len(marine_locations), effective_batch)
        for i in range(0, len(marine_locations), effective_batch):
            batch = marine_locations[i:i + effective_batch]
            lats = [loc[0] for loc in batch]
            lons = [loc[1] for loc in batch]
            marine_batch = None
            for retry_idx in range(max_retries):
                try:
                    marine_batch = fetch_marine_data_hourly(lats, lons, 1)
                    break
                except OpenMeteoRateLimitError as exc:
                    retry_after = exc.retry_after if exc.retry_after is not None else min(30.0, 2 ** retry_idx)
                    logger.warning('build_weather_info: marine fallback rate-limited retry=%s/%s sleep=%.1fs', retry_idx + 1, max_retries, float(retry_after))
                    time.sleep(float(retry_after))
                except Exception as exc:
                    logger.warning('build_weather_info: marine fallback batch failed error=%s', exc)
                    break
            for j, loc in enumerate(batch):
                slot = {}
                if marine_batch and j < len(marine_batch) and marine_batch[j]:
                    slot = dict(marine_batch[j][0])
                marine_fallback_map[loc] = slot
                logger.info('open-meteo marine fallback  lat=%.3f lon=%.3f  wave_height=%s wave_dir=%s curr_vel=%s curr_dir=%s', loc[0], loc[1], slot.get('wave_height'), slot.get('wave_direction'), slot.get('ocean_current_velocity'), slot.get('ocean_current_direction'))
    return (weather_fallback_map, marine_fallback_map)
def build_weather_info_from_context(path_latlon, weather_context):
    weather_info_list = []
    if not path_latlon:
        return weather_info_list
    sampled_points = []
    weather_patch_locations = []
    marine_patch_locations = []
    cumulative_km = 0.0
    for idx, (lat, lon) in enumerate(path_latlon):
        if idx > 0:
            prev_lat, prev_lon = path_latlon[idx - 1]
            cumulative_km += haversine_distance((prev_lon, prev_lat), (lon, lat)) * 1.852
        eta_hour = int(cumulative_km / max(ROUTING_BASE_SPEED_KMH, 0.1))
        current_weather, current_marine = interpolate_weather_from_context(weather_context, float(lat), float(lon), eta_hour)
        cw_fields = current_weather.get('current', {})
        cm_fields = current_marine.get('current', {})
        lat_q, lon_q = _quantize_location(float(lat), float(lon), precision=3)
        loc_key = (lat_q, lon_q)
        missing_weather_fields, missing_marine_fields = _missing_field_names(cw_fields, cm_fields)
        if missing_weather_fields:
            weather_patch_locations.append(loc_key)
        if missing_marine_fields:
            marine_patch_locations.append(loc_key)
        sampled_points.append({'lat': float(lat), 'lon': float(lon), 'loc_key': loc_key, 'cw_fields': dict(cw_fields), 'cm_fields': dict(cm_fields)})
    weather_fallback_map = {}
    marine_fallback_map = {}
    if weather_patch_locations or marine_patch_locations:
        weather_fallback_map, marine_fallback_map = _batch_openmeteo_fallback_grouped(weather_patch_locations, marine_patch_locations, batch_size=OPEN_METEO_BATCH_SIZE)
    for item in sampled_points:
        cw_fields = item['cw_fields']
        cm_fields = item['cm_fields']
        loc_key = item['loc_key']
        fb_cw = weather_fallback_map.get(loc_key, {})
        fb_cm = marine_fallback_map.get(loc_key, {})
        cw_fields, cm_fields = _patch_with_fallback(cw_fields, cm_fields, fb_cw, fb_cm)
        risk_score = round(compute_safety_risk(cw_fields, cm_fields) * 100.0, 1)
        weather_info_list.append({'coordinate': [item['lat'], item['lon']], 'wind_speed': cw_fields.get('wind_speed_10m'), 'wind_direction': cw_fields.get('wind_direction_10m'), 'precipitation': cw_fields.get('precipitation'), 'visibility': cw_fields.get('visibility'), 'wave_height': cm_fields.get('wave_height'), 'wave_dir': cm_fields.get('wave_direction'), 'current_vel': cm_fields.get('ocean_current_velocity'), 'current_dir': cm_fields.get('ocean_current_direction'), 'risk': risk_score})
    _FILL_FIELDS = {'wave_height': True, 'wave_dir': False, 'wind_speed': True, 'wind_direction': False, 'visibility': True, 'current_vel': False, 'current_dir': False, 'precipitation': False}
    for field, zero_invalid in _FILL_FIELDS.items():
        valid_vals = [float(p[field]) for p in weather_info_list if p.get(field) is not None and np.isfinite(float(p[field])) and (not (zero_invalid and float(p[field]) == 0.0))]
        if not valid_vals:
            continue
        avg = sum(valid_vals) / len(valid_vals)
        filled = 0
        for p in weather_info_list:
            v = p.get(field)
            is_bad = v is None or not np.isfinite(float(v)) or (zero_invalid and float(v) == 0.0)
            if is_bad:
                p[field] = round(avg, 3)
                filled += 1
        if filled:
            logger.info('route-avg fill  field=%s  avg=%.3f  filled=%d/%d points', field, avg, filled, len(weather_info_list))
    return weather_info_list
def _adaptive_edge_sample_count(u_node, v_node):
    if not ADAPTIVE_EDGE_SAMPLING_ENABLED:
        return max(2, int(EDGE_WEATHER_SAMPLES))
    edge_km = haversine_distance((u_node[0], u_node[1]), (v_node[0], v_node[1])) * 1.852
    if edge_km <= 25.0:
        return 2
    if edge_km <= 80.0:
        return 3
    if edge_km <= 180.0:
        return 4
    return 5
def interpolate_edge_samples(u_node, v_node, sample_count):
    sample_count = max(2, int(sample_count))
    u_lat, u_lon = (float(u_node[1]), float(u_node[0]))
    v_lat, v_lon = (float(v_node[1]), float(v_node[0]))
    points = []
    for i in range(sample_count):
        t = i / (sample_count - 1)
        lat = u_lat + t * (v_lat - u_lat)
        lon = u_lon + t * (v_lon - u_lon)
        points.append(_quantize_location(lat, lon))
    return points
ROUTING_BASE_SPEED_KMH = 14.0 * 1.852
async def batch_fetch_weather_forecast(locations, forecast_hours=72, batch_size=50):
    hourly_weather: dict = {}
    hourly_marine: dict = {}
    streaming_severity: list = []
    unique_locations = list(dict.fromkeys(locations))
    if not unique_locations:
        return (hourly_weather, hourly_marine, streaming_severity)
    total_batches = (len(unique_locations) + batch_size - 1) // batch_size
    est_http_calls = total_batches * 2
    start_ts = time.perf_counter()
    logger.info('[weather-forecast] points=%s hours=%s batch_size=%s batches=%s est_http_calls=%s', len(unique_locations), forecast_hours, batch_size, total_batches, est_http_calls)
    _last_batch_end = time.perf_counter()
    MIN_BATCH_INTERVAL_S = 1.5
    for i in range(0, len(unique_locations), batch_size):
        batch = unique_locations[i:i + batch_size]
        lats = [loc[0] for loc in batch]
        lons = [loc[1] for loc in batch]
        weather_batch = None
        marine_batch = None
        for retry_idx in range(OPEN_METEO_MAX_RETRIES):
            try:
                await RATE_LIMITER.wait_for_slot(units=2)
                if i > 0:
                    elapsed_since_last = time.perf_counter() - _last_batch_end
                    gap = MIN_BATCH_INTERVAL_S - elapsed_since_last
                    if gap > 0:
                        await asyncio.sleep(gap)
                weather_batch, marine_batch = await asyncio.gather(asyncio.to_thread(fetch_weather_data_hourly, lats, lons, forecast_hours), asyncio.to_thread(fetch_marine_data_hourly, lats, lons, forecast_hours))
                break
            except OpenMeteoRateLimitError as exc:
                RATE_LIMITER.on_rate_limit(retry_idx, retry_after=exc.retry_after)
        fallback_w = [{'wind_speed_10m': 10.0, 'wind_direction_10m': 180.0}] * forecast_hours
        fallback_m = [{'wave_height': 1.0, 'wave_direction': 180.0, 'ocean_current_velocity': 2.0, 'ocean_current_direction': 180.0}] * forecast_hours
        for j, loc in enumerate(batch):
            hourly_weather[loc] = weather_batch[j] if weather_batch and j < len(weather_batch) else fallback_w
            hourly_marine[loc] = marine_batch[j] if marine_batch and j < len(marine_batch) else fallback_m
            cw = hourly_weather[loc][0] if hourly_weather[loc] else {}
            cm = hourly_marine[loc][0] if hourly_marine[loc] else {}
            streaming_severity.append({'coordinate': [loc[0], loc[1]], 'risk': round(compute_safety_risk(cw, cm) * 100.0, 2)})
        _last_batch_end = time.perf_counter()
        elapsed = max(_last_batch_end - start_ts, 1e-06)
        done_batches = min(i // batch_size + 1, total_batches)
        batches_per_sec = done_batches / elapsed
        eta_seconds = max(total_batches - done_batches, 0) / max(batches_per_sec, 1e-06)
        logger.info('[%s] Forecast batch %s/%s (%.1f%%) elapsed=%.1fs eta=%.1fs', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), done_batches, total_batches, done_batches / total_batches * 100.0, elapsed, eta_seconds)
    total_elapsed = time.perf_counter() - start_ts
    logger.info('[weather-forecast] completed points=%s hours=%s batches=%s total_elapsed=%.1fs severity_pts=%s', len(unique_locations), forecast_hours, total_batches, total_elapsed, len(streaming_severity))
    return (hourly_weather, hourly_marine, streaming_severity)
def _is_valid_number(value):
    try:
        f = float(value)
        return np.isfinite(f)
    except (TypeError, ValueError):
        return False
def _circular_mean_degrees(values):
    radians_list = [np.radians(v) for v in values]
    sin_mean = np.mean(np.sin(radians_list))
    cos_mean = np.mean(np.cos(radians_list))
    return (np.degrees(np.arctan2(sin_mean, cos_mean)) + 360.0) % 360.0
def _nearest_locations_with_field(target_loc, candidates, field_name, field_type):
    scored = []
    for loc, payload in candidates:
        current = payload.get('current', {})
        value = current.get(field_name)
        if not _is_valid_number(value):
            continue
        d = haversine_distance((target_loc[1], target_loc[0]), (loc[1], loc[0]))
        scored.append((d, float(value)))
    scored.sort(key=lambda x: x[0])
    top = [v for _, v in scored[:4]]
    if not top:
        return None
    if field_type == 'angle':
        return _circular_mean_degrees(top)
    return float(np.mean(top))
def impute_missing_weather_from_neighbors(locations, weather_results, marine_results):
    weather_fields = [('wind_speed_10m', 'scalar'), ('wind_direction_10m', 'angle')]
    marine_fields = [('wave_height', 'scalar'), ('wave_direction', 'angle'), ('ocean_current_velocity', 'scalar'), ('ocean_current_direction', 'angle')]
    weather_candidates = [(loc, weather_results.get(loc, {})) for loc in locations]
    marine_candidates = [(loc, marine_results.get(loc, {})) for loc in locations]
    for loc in locations:
        weather_results.setdefault(loc, {'current': {}})
        marine_results.setdefault(loc, {'current': {}})
        weather_results[loc].setdefault('current', {})
        marine_results[loc].setdefault('current', {})
        for field_name, field_type in weather_fields:
            if not _is_valid_number(weather_results[loc]['current'].get(field_name)):
                imputed = _nearest_locations_with_field(loc, weather_candidates, field_name, field_type)
                if imputed is not None:
                    weather_results[loc]['current'][field_name] = imputed
        for field_name, field_type in marine_fields:
            if not _is_valid_number(marine_results[loc]['current'].get(field_name)):
                imputed = _nearest_locations_with_field(loc, marine_candidates, field_name, field_type)
                if imputed is not None:
                    marine_results[loc]['current'][field_name] = imputed
    return (weather_results, marine_results)
def summarize_route_metrics(label, path_km, weather_points):
    wave_values = []
    wind_values = []
    precip_values = []
    vis_values = []
    for point in weather_points:
        w = _to_float_or_none(point.get('wave_height'))
        s = _to_float_or_none(point.get('wind_speed'))
        p = _to_float_or_none(point.get('precipitation'))
        v = _to_float_or_none(point.get('visibility'))
        if w is not None:
            wave_values.append(w)
        if s is not None:
            wind_values.append(s)
        if p is not None:
            precip_values.append(p)
        if v is not None:
            vis_values.append(v)
    avg_wave = round(sum(wave_values) / len(wave_values), 3) if wave_values else None
    max_wave = round(max(wave_values), 3) if wave_values else None
    avg_wind = round(sum(wind_values) / len(wind_values), 3) if wind_values else None
    avg_precip = round(sum(precip_values) / len(precip_values), 3) if precip_values else None
    avg_vis = round(sum(vis_values) / len(vis_values), 1) if vis_values else None
    total_points = max(len(weather_points), 1)
    coverage = max(len(wave_values), len(wind_values)) / total_points
    synthetic_weather = {'wind_speed_10m': avg_wind or 0.0, 'precipitation': avg_precip or 0.0, 'visibility': avg_vis or 10000.0}
    synthetic_marine = {'wave_height': avg_wave or 0.0}
    if wave_values or wind_values:
        risk_score = round(compute_safety_risk(synthetic_weather, synthetic_marine) * 100.0, 2)
    else:
        risk_score = 50.0
    return {'label': label, 'distance_km': round(path_km, 3), 'avg_wave_m': avg_wave, 'max_wave_m': max_wave, 'avg_wind_kmh': avg_wind, 'avg_precipitation_mm': avg_precip, 'avg_visibility_m': avg_vis, 'risk_score': risk_score, 'weather_coverage_pct': round(coverage * 100.0, 2)}
def compute_point_risk_score(current_weather, current_marine):
    cw = current_weather.get('current', {})
    cm = current_marine.get('current', {})
    return round(compute_safety_risk(cw, cm) * 100.0, 2)
def build_severity_points_from_lookups(locations, weather_lookup, marine_lookup, max_points=1500):
    severity_points = []
    for loc in locations[:max_points]:
        lat, lon = loc
        current_weather = weather_lookup.get(loc, {})
        current_marine = marine_lookup.get(loc, {})
        risk = compute_point_risk_score(current_weather, current_marine)
        severity_points.append({'coordinate': [lat, lon], 'risk': risk})
    return severity_points
def _speed_dir_to_uv(speed, direction_deg):
    if speed is None or direction_deg is None:
        return (None, None)
    rad = np.radians(direction_deg)
    u = float(speed) * float(np.sin(rad))
    v = float(speed) * float(np.cos(rad))
    return (u, v)
def _idw_at_point(sample_x, sample_y, sample_vals, x, y, k=8, power=2):
    if len(sample_vals) == 0:
        return 0.0
    dx = sample_x - x
    dy = sample_y - y
    dist = np.sqrt(dx * dx + dy * dy)
    nearest = np.argmin(dist)
    if dist[nearest] < 1e-09:
        return float(sample_vals[nearest])
    k = min(k, len(sample_vals))
    idx = np.argpartition(dist, k - 1)[:k]
    d = np.maximum(dist[idx], 1e-06)
    weights = 1.0 / d ** power
    return float(np.sum(weights * sample_vals[idx]) / np.sum(weights))
def build_vector_grid_from_lookups(locations, weather_lookup, marine_lookup, grid_rows=64, grid_cols=64):
    points_lon = []
    points_lat = []
    wind_u = []
    wind_v = []
    current_u = []
    current_v = []
    for lat, lon in locations:
        current_weather = weather_lookup.get((lat, lon), {}).get('current', {})
        current_marine = marine_lookup.get((lat, lon), {}).get('current', {})
        ws = _to_float_or_none(current_weather.get('wind_speed_10m'))
        wd = _to_float_or_none(current_weather.get('wind_direction_10m'))
        cu_speed = _to_float_or_none(current_marine.get('ocean_current_velocity'))
        cu_dir = _to_float_or_none(current_marine.get('ocean_current_direction'))
        wu, wv = _speed_dir_to_uv(ws, wd) if ws is not None and wd is not None else (None, None)
        cu, cv = _speed_dir_to_uv(cu_speed, cu_dir) if cu_speed is not None and cu_dir is not None else (None, None)
        if wu is None and cu is None:
            continue
        points_lon.append(float(lon))
        points_lat.append(float(lat))
        wind_u.append(float(wu or 0.0))
        wind_v.append(float(wv or 0.0))
        current_u.append(float(cu or 0.0))
        current_v.append(float(cv or 0.0))
    if len(points_lon) < 3:
        return None
    min_lon = min(points_lon) - 0.3
    max_lon = max(points_lon) + 0.3
    min_lat = min(points_lat) - 0.3
    max_lat = max(points_lat) + 0.3
    gx = np.linspace(min_lon, max_lon, grid_cols)
    gy = np.linspace(min_lat, max_lat, grid_rows)
    sx = np.array(points_lon)
    sy = np.array(points_lat)
    swu = np.array(wind_u)
    swv = np.array(wind_v)
    scu = np.array(current_u)
    scv = np.array(current_v)
    out_wu = np.zeros((grid_rows, grid_cols), dtype=np.float32)
    out_wv = np.zeros((grid_rows, grid_cols), dtype=np.float32)
    out_cu = np.zeros((grid_rows, grid_cols), dtype=np.float32)
    out_cv = np.zeros((grid_rows, grid_cols), dtype=np.float32)
    for r, lat in enumerate(gy):
        for c, lon in enumerate(gx):
            out_wu[r, c] = _idw_at_point(sx, sy, swu, lon, lat, k=8)
            out_wv[r, c] = _idw_at_point(sx, sy, swv, lon, lat, k=8)
            out_cu[r, c] = _idw_at_point(sx, sy, scu, lon, lat, k=8)
            out_cv[r, c] = _idw_at_point(sx, sy, scv, lon, lat, k=8)
    return {'bounds': {'min_lon': round(min_lon, 6), 'max_lon': round(max_lon, 6), 'min_lat': round(min_lat, 6), 'max_lat': round(max_lat, 6)}, 'rows': grid_rows, 'cols': grid_cols, 'wind_u': out_wu.ravel().tolist(), 'wind_v': out_wv.ravel().tolist(), 'current_u': out_cu.ravel().tolist(), 'current_v': out_cv.ravel().tolist()}
def build_severity_grid_payload(severity_points, rows=SEVERITY_GRID_ROWS, cols=SEVERITY_GRID_COLS):
    if not severity_points:
        return None
    points = []
    values = []
    for point in severity_points:
        coord = point.get('coordinate') if isinstance(point, dict) else None
        if not isinstance(coord, (list, tuple)) or len(coord) != 2:
            continue
        lat = _to_float_or_none(coord[0])
        lon = _to_float_or_none(coord[1])
        risk = _to_float_or_none(point.get('risk') if isinstance(point, dict) else None)
        if lat is None or lon is None or risk is None:
            continue
        points.append((lon, lat))
        values.append(float(np.clip(risk, 0.0, 100.0)))
    if len(points) < 4:
        return None
    points_arr = np.asarray(points, dtype=np.float32)
    values_arr = np.asarray(values, dtype=np.float32)
    min_lon = float(np.min(points_arr[:, 0])) - SEVERITY_GRID_PADDING_DEG
    max_lon = float(np.max(points_arr[:, 0])) + SEVERITY_GRID_PADDING_DEG
    min_lat = float(np.min(points_arr[:, 1])) - SEVERITY_GRID_PADDING_DEG
    max_lat = float(np.max(points_arr[:, 1])) + SEVERITY_GRID_PADDING_DEG
    if max_lon <= min_lon or max_lat <= min_lat:
        return None
    grid_cols = max(int(cols), 24)
    grid_rows = max(int(rows), 24)
    lon_values = np.linspace(min_lon, max_lon, grid_cols, dtype=np.float32)
    lat_values = np.linspace(max_lat, min_lat, grid_rows, dtype=np.float32)
    grid_lon, grid_lat = np.meshgrid(lon_values, lat_values)
    grid = griddata(points_arr, values_arr, (grid_lon, grid_lat), method='linear', fill_value=np.nan)
    if np.any(~np.isfinite(grid)):
        grid_nearest = griddata(points_arr, values_arr, (grid_lon, grid_lat), method='nearest')
        grid = np.where(np.isfinite(grid), grid, grid_nearest)
    grid = np.clip(grid, 0.0, 100.0)
    grid = np.round(grid.astype(np.float32), 2)
    return {'bounds': {'min_lon': round(min_lon, 6), 'max_lon': round(max_lon, 6), 'min_lat': round(min_lat, 6), 'max_lat': round(max_lat, 6)}, 'rows': grid_rows, 'cols': grid_cols, 'values': grid.ravel().tolist()}
def _to_float_or_none(value):
    try:
        number = float(value)
        if np.isfinite(number):
            return number
    except (TypeError, ValueError):
        return None
    return None
def _bearing_from_latlon(point_a, point_b):
    lat1, lon1 = map(radians, point_a)
    lat2, lon2 = map(radians, point_b)
    dlon = lon2 - lon1
    x = sin(dlon) * cos(lat2)
    y = cos(lat1) * sin(lat2) - sin(lat1) * cos(lat2) * cos(dlon)
    return (np.degrees(atan2(x, y)) + 360.0) % 360.0
def estimate_route_fuel_proxy(path_latlon, weather_info):
    if len(path_latlon) < 2:
        return 0.0
    total_fuel_proxy = 0.0
    for i in range(len(path_latlon) - 1):
        current_point = path_latlon[i]
        next_point = path_latlon[i + 1]
        segment_km = haversine_distance((current_point[1], current_point[0]), (next_point[1], next_point[0])) * 1.852
        bearing = _bearing_from_latlon(current_point, next_point)
        wp = weather_info[i] if i < len(weather_info) else {}
        wind_speed = _to_float_or_none(wp.get('wind_speed'))
        wind_dir = _to_float_or_none(wp.get('wind_direction'))
        wave_height = _to_float_or_none(wp.get('wave_height'))
        wave_dir = _to_float_or_none(wp.get('wave_dir'))
        current_vel = _to_float_or_none(wp.get('current_vel'))
        current_dir = _to_float_or_none(wp.get('current_dir'))
        weather_payload = {'weather': {'current': {}}, 'marine': {'current': {}}}
        if wind_speed is not None:
            weather_payload['weather']['current']['wind_speed_10m'] = wind_speed
        if wind_dir is not None:
            weather_payload['weather']['current']['wind_direction_10m'] = wind_dir
        if wave_height is not None:
            weather_payload['marine']['current']['wave_height'] = wave_height
        if wave_dir is not None:
            weather_payload['marine']['current']['wave_direction'] = wave_dir
        if current_vel is not None:
            weather_payload['marine']['current']['ocean_current_velocity'] = current_vel
        if current_dir is not None:
            weather_payload['marine']['current']['ocean_current_direction'] = current_dir
        weather_penalty = calculate_weather_cost(weather_payload, bearing)
        bounded_penalty = min(max(float(weather_penalty), 0.0), 2.0)
        total_fuel_proxy += segment_km * (1.0 + 0.35 * bounded_penalty)
    return round(total_fuel_proxy, 3)
def build_mode_explanation(mode, profile, astar_metrics, optimized_metrics, fuel_saved, fuel_saved_percent, edge_diagnostics=None):
    reasons = []
    risk_a = astar_metrics.get('risk_score')
    risk_o = optimized_metrics.get('risk_score')
    distance_saved_km = round(astar_metrics['distance_km'] - optimized_metrics['distance_km'], 3)
    reasons.append(f'distance_delta_km={distance_saved_km}')
    if risk_a is not None and risk_o is not None:
        reasons.append(f'risk_delta={round(risk_a - risk_o, 2)}')
    reasons.append(f'fuel_proxy_saved={fuel_saved}')
    focus = {'safety': 'Prioritizes lower weather risk (wind/wave/current + direction alignment).', 'distance': 'Prioritizes shortest baseline route (A*).', 'fuel consumption': 'Prioritizes lower directional weather resistance to reduce fuel burn.', 'balanced': 'Minimizes voyage economics (time charter + fuel) with risk penalty only above a safety band.'}.get(mode, 'Balanced objective.')
    return {'mode': mode, 'focus': focus, 'weights': profile, 'vessel_limits': {'max_safe_wave_height_m': MAX_SAFE_WAVE_HEIGHT_M, 'max_safe_wind_kmh': MAX_SAFE_WIND_KMH, 'max_safe_current_kmh': MAX_SAFE_CURRENT_KMH}, 'factors_considered': ['edge_distance', 'edge_time_hours', 'fuel_tonnes', 'wind_speed_direction', 'wave_height_direction', 'ocean_current_velocity_direction'], 'edge_diagnostics': edge_diagnostics or {}, 'why_optimized': reasons, 'fuel_saved_proxy': fuel_saved, 'fuel_saved_proxy_percent': fuel_saved_percent}
def sanitize_for_json(value):
    if isinstance(value, np.ndarray):
        flat = value.flatten().tolist()
        flat = [None if isinstance(x, float) and (not x == x) else x for x in flat]
        if value.ndim == 1:
            return flat
        cols = value.shape[1]
        return [flat[i:i + cols] for i in range(0, len(flat), cols)]
    if isinstance(value, dict):
        return {k: sanitize_for_json(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        if value and all((type(x) in (int, float, str, bool, type(None)) for x in value)):
            return [None if isinstance(x, float) and (not x == x) else x for x in value]
        return [sanitize_for_json(v) for v in value]
    if isinstance(value, np.generic):
        v = value.item()
        return None if isinstance(v, float) and (not v == v) else v
    if isinstance(value, float):
        return None if not value == value or value == float('inf') or value == float('-inf') else value
    if isinstance(value, (int, str, bool)) or value is None:
        return value
    return str(value)
def is_no_go_weather(current_weather, current_marine):
    wind_speed = _to_float_or_none(current_weather.get('current', {}).get('wind_speed_10m'))
    wave_height = _to_float_or_none(current_marine.get('current', {}).get('wave_height'))
    current_vel = _to_float_or_none(current_marine.get('current', {}).get('ocean_current_velocity'))
    if wave_height is not None and wave_height > MAX_SAFE_WAVE_HEIGHT_M:
        return True
    if wind_speed is not None and wind_speed > MAX_SAFE_WIND_KMH:
        return True
    if current_vel is not None and current_vel > MAX_SAFE_CURRENT_KMH:
        return True
    return False
def is_missing_weather(current_weather, current_marine):
    wind_speed = _to_float_or_none(current_weather.get('current', {}).get('wind_speed_10m'))
    wave_height = _to_float_or_none(current_marine.get('current', {}).get('wave_height'))
    return wind_speed is None or wave_height is None
def aggregate_edge_weather(sample_payloads, bearing):
    penalties = []
    has_missing = False
    no_go_detected = False
    for current_weather, current_marine in sample_payloads:
        if is_missing_weather(current_weather, current_marine):
            has_missing = True
        if is_no_go_weather(current_weather, current_marine):
            no_go_detected = True
        penalty = calculate_weather_cost({'weather': current_weather, 'marine': current_marine}, bearing)
        if np.isfinite(penalty):
            penalties.append(float(penalty))
    if not penalties:
        return (1.0, has_missing, no_go_detected)
    mean_penalty = float(np.mean(penalties))
    max_penalty = float(np.max(penalties))
    aggregated_penalty = 0.35 * mean_penalty + 0.65 * max_penalty
    return (aggregated_penalty, has_missing, no_go_detected)
def compute_path_cost(graph, path, weight_key='weight'):
    if not path or len(path) < 2:
        return 0.0
    total = 0.0
    for i in range(len(path) - 1):
        u = path[i]
        v = path[i + 1]
        edge_data = graph.get_edge_data(u, v, default={})
        w = edge_data.get(weight_key)
        if w is None:
            w = edge_data.get('distance', edge_data.get('weight', 0.0))
        try:
            total += float(w)
        except (TypeError, ValueError):
            total += 0.0
    return round(total, 6)
def _edge_weight_value(graph, u, v, weight_key='weight'):
    edge_data = graph.get_edge_data(u, v, default={})
    value = edge_data.get(weight_key)
    if value is None:
        value = edge_data.get('distance', edge_data.get('weight', 0.0))
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
def _turn_angle_between_edges(prev_node, curr_node, next_node):
    b1 = _calculate_geographic_bearing(prev_node, curr_node)
    b2 = _calculate_geographic_bearing(curr_node, next_node)
    delta = abs(b2 - b1) % 360.0
    return min(delta, 360.0 - delta)
def find_turn_aware_path(graph, start_node, end_node, edge_weight_key='weight', turn_penalty=0.0, max_turn_deg=MAX_ROUTE_TURN_DEG):
    if start_node == end_node:
        return [start_node]
    if start_node not in graph or end_node not in graph:
        return None
    max_turn = float(max_turn_deg)
    turn_pen = max(0.0, float(turn_penalty))
    def get_h(node):
        dist_nm = haversine_distance(node, end_node)
        return dist_nm / max(BASE_SPEED_KNOTS, 1.0) * TIME_CHARTER_RATE_PER_HOUR
    frontier = []
    best_g = {}
    parent = {}
    best_end_state = None
    best_end_cost = float('inf')
    for nbr in graph.neighbors(start_node):
        g = _edge_weight_value(graph, start_node, nbr, edge_weight_key)
        h = get_h(nbr)
        state = (start_node, nbr)
        best_g[state] = g
        parent[state] = None
        heapq.heappush(frontier, (g + h, g, state))
    while frontier:
        f, g_so_far, state = heapq.heappop(frontier)
        if g_so_far > best_g.get(state, float('inf')):
            continue
        prev_node, curr_node = state
        if curr_node == end_node:
            best_end_state = state
            best_end_cost = g_so_far
            break
        for nxt in graph.neighbors(curr_node):
            if nxt == prev_node:
                continue
            turn_angle = _turn_angle_between_edges(prev_node, curr_node, nxt)
            if turn_angle > max_turn:
                continue
            turn_cost = turn_pen * (turn_angle / max(max_turn, 1e-06)) ** 2
            edge_cost = _edge_weight_value(graph, curr_node, nxt, edge_weight_key)
            new_g = g_so_far + edge_cost + turn_cost
            next_state = (curr_node, nxt)
            if new_g < best_g.get(next_state, float('inf')):
                best_g[next_state] = new_g
                parent[next_state] = state
                h = get_h(nxt)
                heapq.heappush(frontier, (new_g + h, new_g, next_state))
    if best_end_state is None:
        return None
    reversed_path = [best_end_state[1], best_end_state[0]]
    cursor = parent.get(best_end_state)
    while cursor is not None:
        reversed_path.append(cursor[0])
        cursor = parent.get(cursor)
    return list(reversed(reversed_path))
def _catmull_rom_spline(P0, P1, P2, P3, num_points):
    alpha = 0.5
    def get_t(t, p0, p1):
        d = sqrt((p1[0] - p0[0]) ** 2 + (p1[1] - p0[1]) ** 2)
        return t + d ** alpha if d > 0 else t + 1e-06
    t0 = 0.0
    t1 = get_t(t0, P0, P1)
    t2 = get_t(t1, P1, P2)
    t3 = get_t(t2, P2, P3)
    t = np.linspace(t1, t2, num_points)
    t = t[:, np.newaxis]
    A1 = (t1 - t) / (t1 - t0) * np.array(P0) + (t - t0) / (t1 - t0) * np.array(P1)
    A2 = (t2 - t) / (t2 - t1) * np.array(P1) + (t - t1) / (t2 - t1) * np.array(P2)
    A3 = (t3 - t) / (t3 - t2) * np.array(P2) + (t - t2) / (t3 - t2) * np.array(P3)
    B1 = (t2 - t) / (t2 - t0) * A1 + (t - t0) / (t2 - t0) * A2
    B2 = (t3 - t) / (t3 - t1) * A2 + (t - t1) / (t3 - t1) * A3
    C = (t2 - t) / (t2 - t1) * B1 + (t - t1) / (t2 - t1) * B2
    return C.tolist()
def constrained_smooth_path(path, points_per_edge=CONSTRAINED_SMOOTH_POINTS_PER_EDGE, max_points=CONSTRAINED_SMOOTH_MAX_POINTS):
    if not path or len(path) < 2:
        return list(path or [])
    if len(path) < 4:
        step_points = max(1, int(points_per_edge))
        dense = [path[0]]
        for i in range(len(path) - 1):
            u, v = path[i], path[i+1]
            for j in range(1, step_points + 1):
                t = j / float(step_points)
                lon = float(u[0]) + t * (float(v[0]) - float(u[0]))
                lat = float(u[1]) + t * (float(v[1]) - float(u[1]))
                dense.append((lon, lat))
        return dense[:max_points]
    dense = [path[0]]
    ext_path = [path[0]] + list(path) + [path[-1]]
    step_points = max(2, int(points_per_edge))
    for i in range(1, len(ext_path) - 2):
        P0, P1, P2, P3 = ext_path[i-1], ext_path[i], ext_path[i+1], ext_path[i+2]
        pts = _catmull_rom_spline(P0, P1, P2, P3, step_points + 1)
        dense.extend([(p[0], p[1]) for p in pts[1:]])
    if len(dense) <= max_points:
        return dense
    stride = max(1, len(dense) // max_points)
    reduced = dense[::stride][:max_points]
    if reduced[-1] != dense[-1]:
        reduced.append(dense[-1])
    return reduced
def _interpolate_latlon_segment(start_latlon, end_latlon, max_step_km=ENDPOINT_CONNECTOR_MAX_STEP_KM):
    s_lat, s_lon = (float(start_latlon[0]), float(start_latlon[1]))
    e_lat, e_lon = (float(end_latlon[0]), float(end_latlon[1]))
    segment_km = haversine_distance((s_lon, s_lat), (e_lon, e_lat)) * 1.852
    if not np.isfinite(segment_km) or segment_km <= 0:
        return []
    n_steps = max(1, int(np.ceil(segment_km / max(float(max_step_km), 0.1))))
    out = []
    for i in range(1, n_steps):
        t = i / float(n_steps)
        lat = s_lat + t * (e_lat - s_lat)
        lon = s_lon + t * (e_lon - s_lon)
        out.append((lat, lon))
    return out
def attach_endpoint_connectors(path_latlon, start_latlon, end_latlon):
    if not path_latlon:
        return [start_latlon, end_latlon]
    start_ll = (float(start_latlon[0]), float(start_latlon[1]))
    end_ll = (float(end_latlon[0]), float(end_latlon[1]))
    first = (float(path_latlon[0][0]), float(path_latlon[0][1]))
    last = (float(path_latlon[-1][0]), float(path_latlon[-1][1]))
    out = [start_ll]
    out.extend(_interpolate_latlon_segment(start_ll, first))
    out.extend(path_latlon)
    out.extend(_interpolate_latlon_segment(last, end_ll))
    out.append(end_ll)
    deduped = []
    for lat, lon in out:
        if not deduped:
            deduped.append((lat, lon))
            continue
        prev_lat, prev_lon = deduped[-1]
        if abs(prev_lat - lat) < 1e-08 and abs(prev_lon - lon) < 1e-08:
            continue
        deduped.append((lat, lon))
    return deduped
def _path_distance_km(graph, path):
    if not path or len(path) < 2:
        return float('inf')
    total = 0.0
    for i in range(len(path) - 1):
        u, v = (path[i], path[i + 1])
        edge_data = graph.get_edge_data(u, v, default={})
        d = edge_data.get('distance', edge_data.get('weight'))
        try:
            total += float(d)
        except (TypeError, ValueError):
            total += 0.0
    return total
def select_best_start_end_nodes(graph, tree, start_coord, end_coord, heuristic_fn):
    k = max(1, int(ENDPOINT_SNAP_TOP_K))
    start_candidates = find_k_nearest_water_nodes(graph, start_coord, tree, k=k)
    end_candidates = find_k_nearest_water_nodes(graph, end_coord, tree, k=k)
    if not start_candidates or not end_candidates:
        start_node = find_nearest_water_node(graph, start_coord, tree)
        end_node = find_nearest_water_node(graph, end_coord, tree)
        return (start_node, end_node, None)
    candidate_pairs = []
    for s_node, s_dist in start_candidates:
        for e_node, e_dist in end_candidates:
            candidate_pairs.append((s_node, e_node, s_dist + e_dist))
    candidate_pairs.sort(key=lambda item: item[2])
    limit = min(len(candidate_pairs), max(1, int(ENDPOINT_PAIR_MAX_COMBINATIONS)))
    best_pair = None
    for s_node, e_node, _ in candidate_pairs[:limit]:
        try:
            trial_path = nx.astar_path(graph, s_node, e_node, heuristic=heuristic_fn, weight='distance')
        except nx.NetworkXNoPath:
            continue
        trial_cost = _path_distance_km(graph, trial_path)
        if best_pair is None or trial_cost < best_pair[3]:
            best_pair = (s_node, e_node, trial_path, trial_cost)
    if best_pair is not None:
        return (best_pair[0], best_pair[1], best_pair[2])
    return (start_candidates[0][0], end_candidates[0][0], None)
def count_route_limit_violations(weather_points):
    violations = {'wave': 0, 'wind': 0, 'current': 0}
    for point in weather_points:
        wave = _to_float_or_none(point.get('wave_height'))
        wind = _to_float_or_none(point.get('wind_speed'))
        current_vel = _to_float_or_none(point.get('current_vel'))
        if wave is not None and wave > MAX_SAFE_WAVE_HEIGHT_M:
            violations['wave'] += 1
        if wind is not None and wind > MAX_SAFE_WIND_KMH:
            violations['wind'] += 1
        if current_vel is not None and current_vel > MAX_SAFE_CURRENT_KMH:
            violations['current'] += 1
    violations['total'] = violations['wave'] + violations['wind'] + violations['current']
    return violations
def compute_arrival_hours_from_path(path, graph):
    if not path:
        return {}
    hours = {path[0]: 0.0}
    cumulative = 0.0
    for i in range(len(path) - 1):
        u = path[i]
        v = path[i + 1]
        edge_data = graph.get_edge_data(u, v, default={})
        segment_hours = _to_float_or_none(edge_data.get('time_hours'))
        if segment_hours is None:
            distance_nm = _to_float_or_none(edge_data.get('original_weight', edge_data.get('weight', 1.0))) or 1.0
            effective_sf = _to_float_or_none(edge_data.get('effective_speed_factor')) or 0.7
            speed_knots = max(BASE_SPEED_KNOTS * max(effective_sf, 0.05), 0.1)
            segment_hours = distance_nm / speed_knots
        cumulative += max(float(segment_hours), 0.0)
        hours[v] = cumulative
    return hours
def optimize_path_with_iterative_refinement(subgraph, start_node, end_node, objective_profile, selected_mode, total_distance_km, weather_context, initial_arrival_hours, initial_path, max_iter=OPTIMIZATION_MAX_ITER, convergence_threshold=OPTIMIZATION_CONVERGENCE_THRESHOLD):
    current_arrival_hours = dict(initial_arrival_hours or {})
    current_path = list(initial_path) if initial_path else []
    prev_cost = None
    optimized_subgraph = subgraph.copy()
    edge_diagnostics = {}
    severity_points = weather_context.get('severity_points', [])
    best_path = current_path
    best_cost = float('inf')
    best_arrival_hours = current_arrival_hours
    best_subgraph = optimized_subgraph
    best_edge_diagnostics = edge_diagnostics
    best_severity_points = severity_points
    best_weather_context = weather_context
    for iteration in range(max(1, int(max_iter))):
        optimized_subgraph, edge_diagnostics, severity_points, weather_context = update_subgraph_weights(subgraph.copy(), start_node, end_node, objective_profile, selected_mode, total_distance_km=total_distance_km, weather_context=weather_context, arrival_hours=current_arrival_hours)
        if iteration == 0 and initial_path:
            initial_path_cost = compute_path_cost(optimized_subgraph, initial_path, weight_key='weight')
            logger.info('[iter-opt] mode=%s A*-baseline cost on objective graph=%.2f', selected_mode, initial_path_cost)
            if initial_path_cost < best_cost:
                best_cost = initial_path_cost
                best_path = list(initial_path)
                best_subgraph = optimized_subgraph
                best_edge_diagnostics = edge_diagnostics
                best_severity_points = severity_points
                best_weather_context = weather_context
        new_path = find_turn_aware_path(optimized_subgraph, start_node, end_node, edge_weight_key='weight', turn_penalty=TURN_PENALTY_USD, max_turn_deg=MAX_ROUTE_TURN_DEG)
        if not new_path:
            logger.warning('[iter-opt] no path iteration=%s mode=%s; keeping previous path', iteration + 1, selected_mode)
            break
        new_arrival_hours = compute_arrival_hours_from_path(new_path, optimized_subgraph)
        if current_arrival_hours:
            merged_arrivals = dict(current_arrival_hours)
            merged_arrivals.update(new_arrival_hours)
            new_arrival_hours = merged_arrivals
        new_cost = compute_path_cost(optimized_subgraph, new_path, weight_key='weight')
        gain = (prev_cost - new_cost) if prev_cost is not None else 0.0
        logger.info('[iter-opt] mode=%s iteration=%s cost=%.2f delta=%.2f nodes=%s', selected_mode, iteration + 1, new_cost, gain, len(new_path))
        current_path = new_path
        current_arrival_hours = new_arrival_hours
        if new_cost < best_cost:
            best_cost = new_cost
            best_path = new_path
            best_arrival_hours = new_arrival_hours
            best_subgraph = optimized_subgraph
            best_edge_diagnostics = edge_diagnostics
            best_severity_points = severity_points
            best_weather_context = weather_context
        if prev_cost is not None and prev_cost > 0:
            rel_change = abs(prev_cost - new_cost) / prev_cost
            if rel_change < max(float(convergence_threshold), 0.0):
                logger.info('[iter-opt] converged mode=%s iteration=%s rel_change=%.4f', selected_mode, iteration + 1, rel_change)
                prev_cost = new_cost
                break
        prev_cost = new_cost
    total_gain = initial_path_cost - best_cost if initial_path and 'initial_path_cost' in locals() else 0.0
    logger.info('[iter-opt] best mode=%s final_cost=%.2f total_economic_gain=%.2f', selected_mode, best_cost, total_gain)
    return {'path': best_path if best_path else list(initial_path), 'graph': best_subgraph, 'edge_diagnostics': best_edge_diagnostics, 'severity_points': best_severity_points, 'weather_context': best_weather_context, 'cost': best_cost, 'arrival_hours': best_arrival_hours}
def update_subgraph_weights(subgraph, start_node, end_node, profile, selected_mode, total_distance_km=None, weather_context=None, arrival_hours=None):
    edge_samples = {}
    bearings = {}
    total_sample_points = 0
    for u, v, _ in subgraph.edges(data=True):
        sample_count = _adaptive_edge_sample_count(u, v)
        sample_points = interpolate_edge_samples(u, v, sample_count)
        edge_samples[u, v] = sample_points
        total_sample_points += len(sample_points)
        bearings[u, v] = _calculate_geographic_bearing((u[1], u[0]), (v[1], v[0]))
    if arrival_hours is None:
        arrival_hours = estimate_node_arrival_hours(subgraph, end_node, total_distance_km)
    forecast_hours = min(72, max(24, int(max(arrival_hours.values(), default=0)) + 6))
    if weather_context is None:
        raise ValueError('weather_context is required for update_subgraph_weights')
    if int(weather_context.get('forecast_hours', 0)) < int(forecast_hours):
        logger.warning('[edge-weather] weather_context forecast horizon (%s) shorter than requested (%s)', weather_context.get('forecast_hours', 0), forecast_hours)
    severity_points = weather_context.get('severity_points', [])
    logger.info('[edge-weather] edges=%s sampled_points=%s avg_samples_per_edge=%.2f cube_points=%s', len(edge_samples), total_sample_points, total_sample_points / max(len(edge_samples), 1), len(weather_context.get('grid_points', [])))
    alpha = profile.get('alpha', 1.0)
    beta = profile.get('beta', 1.0)
    gamma = profile.get('gamma', 1.0)
    edge_diagnostics = {'total_edges': len(subgraph.edges()), 'high_weather_edges': 0, 'no_go_edges_blocked': 0, 'missing_weather_edges_blocked': 0, 'edge_sample_points_mode': 'adaptive' if ADAPTIVE_EDGE_SAMPLING_ENABLED else 'fixed', 'avg_edge_samples': round(total_sample_points / max(len(edge_samples), 1), 3), 'weather_cube_points': len(weather_context.get('grid_points', []))}
    for (u, v), sample_points in edge_samples.items():
        try:
            bearing = bearings[u, v]
            data = subgraph[u][v]
            distance_nm = float(data.get('original_weight', data.get('weight', 1.0)))
            data['original_weight'] = distance_nm
            t_start = float(arrival_hours.get(u, 0.0))
            t_end = float(arrival_hours.get(v, t_start))
            sample_payloads = []
            for idx, loc in enumerate(sample_points):
                frac = idx / (len(sample_points) - 1) if len(sample_points) > 1 else 0.0
                t_edge = t_start + frac * (t_end - t_start)
                current_weather, current_marine = interpolate_weather_from_context(weather_context, loc[0], loc[1], int(t_edge))
                sample_payloads.append((current_weather, current_marine))
            weather_penalty, has_missing, no_go_detected = aggregate_edge_weather(sample_payloads, bearing)
            if no_go_detected:
                NO_GO_MULTIPLIER = 50.0
                base_cost = distance_nm * (FUEL_RATE_AT_REF_SPEED_TONNES_PER_NM * FUEL_PRICE_PER_TONNE + TIME_CHARTER_RATE_PER_HOUR / max(BASE_SPEED_KNOTS, 1.0))
                data['weight'] = base_cost * NO_GO_MULTIPLIER
                edge_diagnostics['no_go_edges_blocked'] += 1
                continue
            if selected_mode == 'safety' and SAFETY_BLOCK_ON_MISSING_WEATHER and has_missing:
                NO_GO_MULTIPLIER = 50.0
                base_cost = distance_nm * (FUEL_RATE_AT_REF_SPEED_TONNES_PER_NM * FUEL_PRICE_PER_TONNE + TIME_CHARTER_RATE_PER_HOUR / max(BASE_SPEED_KNOTS, 1.0))
                data['weight'] = base_cost * NO_GO_MULTIPLIER
                edge_diagnostics['missing_weather_edges_blocked'] += 1
                continue
            waves_h, waves_d = [], []
            winds_s, winds_d = [], []
            currs_v, currs_d = [], []
            risk_scores = []
            for cw_pkg, cm_pkg in sample_payloads:
                cw = cw_pkg.get('current', {})
                cm = cm_pkg.get('current', {})
                wh = _to_float_or_none(cm.get('wave_height'))
                wd = _to_float_or_none(cm.get('wave_direction'))
                if wh is not None: waves_h.append(wh)
                if wd is not None: waves_d.append(wd)
                ws = _to_float_or_none(cw.get('wind_speed_10m'))
                wdir = _to_float_or_none(cw.get('wind_direction_10m'))
                if ws is not None: winds_s.append(ws)
                if wdir is not None: winds_d.append(wdir)
                cv = _to_float_or_none(cm.get('ocean_current_velocity'))
                cd = _to_float_or_none(cm.get('ocean_current_direction'))
                if cv is not None: currs_v.append(cv)
                if cd is not None: currs_d.append(cd)
                risk_scores.append(compute_safety_risk(cw, cm))
            avg_wh = float(np.mean(waves_h)) if waves_h else 0.0
            avg_wd = float(np.mean(waves_d)) if waves_d else 0.0
            avg_ws = float(np.mean(winds_s)) if winds_s else 0.0
            avg_wdir = float(np.mean(winds_d)) if winds_d else 0.0
            avg_cv = float(np.mean(currs_v)) if currs_v else 0.0
            avg_cd = float(np.mean(currs_d)) if currs_d else 0.0
            edge_risk = float(np.max(risk_scores)) if risk_scores else 0.0
            from vessel_polar import compute_vessel_performance
            perf = compute_vessel_performance(distance_nm, avg_wh, avg_wd, avg_ws, avg_wdir, avg_cv, avg_cd, bearing)
            time_hours = perf['time_hours']
            fuel_tonnes = perf['fuel_tonnes']
            time_cost_usd = time_hours * TIME_CHARTER_RATE_PER_HOUR
            fuel_cost_usd = fuel_tonnes * FUEL_PRICE_PER_TONNE
            risk_excess = max(edge_risk - RISK_FREE_BAND, 0.0)
            risk_cost_usd = risk_excess ** max(RISK_COST_EXPONENT, 1.0) * distance_nm * RISK_COST_PER_UNIT_RISK
            edge_cost = alpha * time_cost_usd + beta * risk_cost_usd + gamma * fuel_cost_usd
            if selected_mode == 'safety' and edge_risk > 0.25:
                extra = SAFETY_PENALTY_MULTIPLIER * (edge_risk - 0.25) / 0.75
                edge_cost *= 1.0 + extra
            data['weight'] = edge_cost if np.isfinite(edge_cost) else distance_nm
            data['time_hours'] = float(time_hours)
            data['fuel_tonnes'] = float(fuel_tonnes)
            data['risk_cost_usd'] = float(risk_cost_usd)
            data['risk_excess'] = float(risk_excess)
            data['time_cost_usd'] = float(time_cost_usd)
            data['fuel_cost_usd'] = float(fuel_cost_usd)
            data['effective_speed_knots'] = float(perf['speed_kn'])
            data['effective_speed_factor'] = float(perf['effective_sf'])
            data['edge_risk'] = float(edge_risk)
        except Exception as e:
            logger.exception('Error processing edge (%s-%s): %s', u, v, e)
            data = subgraph[u][v]
            data['weight'] = data.get('original_weight', data.get('weight', 1.0))
            data['time_hours'] = data.get('time_hours', 1.0)
    return (subgraph, edge_diagnostics, severity_points, weather_context)
def compute_wind_components(wind_speed, wind_direction):
    rad = math.radians(wind_direction)
    u = -wind_speed * math.sin(rad)
    v = -wind_speed * math.cos(rad)
    return (u, v)
def create_wind_grid(locations, weather_lookup, bbox_padding_deg=0.5, grid_res=0.05):
    if not locations:
        return (None, None, None, None, [], [], [], [])
    points = []
    u_vals = []
    v_vals = []
    for loc in locations:
        w = weather_lookup.get(loc, {}).get('current', {})
        speed = w.get('wind_speed_10m')
        direction = w.get('wind_direction_10m')
        if speed is not None and direction is not None and np.isfinite(speed) and np.isfinite(direction):
            points.append([loc[1], loc[0]])
            u, v = compute_wind_components(speed, direction)
            u_vals.append(u)
            v_vals.append(v)
    if len(points) < 4:
        return (None, None, None, None, [], [], [], [])
    lons = [p[0] for p in points]
    lats = [p[1] for p in points]
    min_lon, max_lon = (min(lons), max(lons))
    min_lat, max_lat = (min(lats), max(lats))
    min_lon -= bbox_padding_deg
    max_lon += bbox_padding_deg
    min_lat -= bbox_padding_deg
    max_lat += bbox_padding_deg
    grid_lons = np.arange(min_lon, max_lon + grid_res, grid_res)
    grid_lats = np.arange(min_lat, max_lat + grid_res, grid_res)
    grid_lon_2d, grid_lat_2d = np.meshgrid(grid_lons, grid_lats)
    points_arr = np.array(points)
    u_arr = np.array(u_vals)
    v_arr = np.array(v_vals)
    method = WIND_GRID_INTERPOLATION_METHOD
    if method not in {'linear', 'nearest', 'cubic'}:
        method = 'linear'
    if method == 'cubic' and len(points) < 16:
        method = 'linear'
    grid_u = griddata(points_arr, u_arr, (grid_lon_2d, grid_lat_2d), method=method, fill_value=np.nan)
    grid_v = griddata(points_arr, v_arr, (grid_lon_2d, grid_lat_2d), method=method, fill_value=np.nan)
    return (min_lon, max_lon, min_lat, max_lat, grid_lons.tolist(), grid_lats.tolist(), grid_u.tolist(), grid_v.tolist())
async def handle_navigation(websocket):
    request_id = str(uuid.uuid4())[:8]
    try:
        if GLOBAL_GRAPH is None or GLOBAL_TREE is None or GLOBAL_NODE_ARRAY is None:
            raise RuntimeError('Navigation graph not initialized. Restart backend.')
        message = await websocket.receive_str()
        data = json.loads(message)
        start_coords = tuple(data['start'])
        end_coords = tuple(data['end'])
        objective_profile, selected_mode = get_objective_profile(data.get('mode', 'balanced'))
        logger.info('request_id=%s start=%s end=%s mode=%s', request_id, start_coords, end_coords, selected_mode)
        start = (start_coords[1], start_coords[0])
        end = (end_coords[1], end_coords[0])
        route_cache_key = (round(float(start[0]), 4), round(float(start[1]), 4), round(float(end[0]), 4), round(float(end[1]), 4), selected_mode, _utc_model_run_key())
        cached_route = ROUTE_RESPONSE_CACHE.get(route_cache_key)
        if cached_route:
            cached_route['created_at'] = time.time()
            logger.info('request_id=%s cache=route-hit mode=%s model_run=%s', request_id, selected_mode, route_cache_key[-1])
            await websocket.send_str(json.dumps(cached_route['payload'], allow_nan=False))
            return
        t_start_nav = time.perf_counter()
        async def _progress(pct, step):
            await websocket.send_str(json.dumps({'type': 'progress', 'pct': pct, 'step': step}))
        async def _await_blocking_with_progress(blocking_fn, start_pct, max_pct, step_messages, tick_seconds=1.5, pct_step=1):
            pct = int(start_pct)
            msg_idx = 0
            task = asyncio.create_task(asyncio.to_thread(blocking_fn))
            while True:
                try:
                    return await asyncio.wait_for(task, timeout=tick_seconds)
                except asyncio.TimeoutError:
                    if pct < max_pct:
                        pct = min(max_pct, pct + pct_step)
                    msg = step_messages[msg_idx % len(step_messages)] if step_messages else 'Working...'
                    await _progress(pct, msg)
                    msg_idx += 1
        def _gc_heuristic_km(u, v):
            from math import radians, sin, cos, atan2, sqrt
            lon1, lat1 = (float(u[0]), float(u[1]))
            lon2, lat2 = (float(v[0]), float(v[1]))
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
            return 2 * 6371.0 * atan2(sqrt(a), sqrt(1 - a))
        await _progress(4, 'Snapping endpoints to navigation graph...')
        t0 = time.perf_counter()
        start_node, end_node, seed_path = await _await_blocking_with_progress(lambda: select_best_start_end_nodes(GLOBAL_GRAPH, GLOBAL_TREE, start, end, heuristic_fn=_gc_heuristic_km), start_pct=4, max_pct=9, step_messages=['Evaluating nearest water-node candidates...', 'Testing start/end snap combinations...', 'Selecting best endpoint pair...'])
        logger.info('request_id=%s endpoint_snap=%.2fs', request_id, time.perf_counter() - t0)
        logger.info('request_id=%s endpoint_snap k=%s start_node=%s end_node=%s seed_path_nodes=%s', request_id, ENDPOINT_SNAP_TOP_K, start_node, end_node, len(seed_path) if seed_path else 0)
        def _geodesic_baseline(start, end, n_waypoints=12):
            from math import radians, degrees, sin, cos, atan2, sqrt
            lon1, lat1 = (radians(float(start[0])), radians(float(start[1])))
            lon2, lat2 = (radians(float(end[0])), radians(float(end[1])))
            waypoints = [start]
            for i in range(1, n_waypoints):
                f = i / n_waypoints
                d = 2 * atan2(sqrt(sin((lat2 - lat1) / 2) ** 2 + cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2) ** 2), sqrt(1 - sin((lat2 - lat1) / 2) ** 2 - cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2) ** 2))
                if d < 1e-09:
                    break
                A = sin((1 - f) * d) / sin(d)
                B = sin(f * d) / sin(d)
                x = A * cos(lat1) * cos(lon1) + B * cos(lat2) * cos(lon2)
                y = A * cos(lat1) * sin(lon1) + B * cos(lat2) * sin(lon2)
                z = A * sin(lat1) + B * sin(lat2)
                lat_i = atan2(z, sqrt(x * x + y * y))
                lon_i = atan2(y, x)
                nearest = find_nearest_water_node(GLOBAL_GRAPH, (degrees(lon_i), degrees(lat_i)), GLOBAL_TREE)
                if nearest not in waypoints:
                    waypoints.append(nearest)
            waypoints.append(end)
            return waypoints
        geodesic_path = _geodesic_baseline(start_node, end_node)
        await _progress(10, 'Finding baseline A* route...')
        if seed_path:
            a_star_path = seed_path
        else:
            try:
                t0 = time.perf_counter()
                a_star_path = await _await_blocking_with_progress(lambda: nx.astar_path(GLOBAL_GRAPH, start_node, end_node, heuristic=_gc_heuristic_km, weight='distance'), start_pct=11, max_pct=18, step_messages=['Running baseline A* search...', 'Expanding open-set nodes...', 'Evaluating route alternatives...'])
                logger.info('request_id=%s baseline_astar=%.2fs', request_id, time.perf_counter() - t0)
            except nx.NetworkXNoPath:
                a_star_path = geodesic_path
        astar_total_distance_km = calculate_total_nautical_distance(a_star_path) * 1.852
        if selected_mode == 'distance':
            await _progress(50, 'Smoothing shortest path...')
            smooth_path = constrained_smooth_path(a_star_path, points_per_edge=CONSTRAINED_SMOOTH_POINTS_PER_EDGE, max_points=CONSTRAINED_SMOOTH_MAX_POINTS)
            path_latlon = [(node[1], node[0]) for node in smooth_path]
            marker_start_latlon = (float(start_coords[0]), float(start_coords[1]))
            marker_end_latlon = (float(end_coords[0]), float(end_coords[1]))
            path_latlon = attach_endpoint_connectors(path_latlon, marker_start_latlon, marker_end_latlon)
            dist_km = calculate_total_nautical_distance(path_latlon) * 1.852
            eta_h = round(dist_km / max(ROUTING_BASE_SPEED_KMH, 0.1), 2)
            metrics_stub = {'label': 'astar', 'distance_km': round(dist_km, 3), 'eta_hours': eta_h, 'eta_days': round(eta_h / 24.0, 2), 'risk_score': None, 'avg_wave_m': None, 'fuel_tonnes': None, 'co2_tonnes': None}
            await _progress(100, 'Done')
            distance_payload = {'type': 'final', 'request_id': request_id, 'path': path_latlon, 'apath': path_latlon, 'weather': [], 'optimized_weather': [], 'astar_weather': [], 'severity': [], 'severity_grid': None, 'wind_field': None, 'wind_grid': None, 'distance': round(dist_km, 3), 'mode': selected_mode, 'excel_export': None, 'alternatives': [], 'metrics': {'astar': metrics_stub, 'optimized': metrics_stub, 'distance_saved_km': 0.0, 'distance_saved_percent': 0.0, 'fuel_saved_proxy': 0.0, 'fuel_saved_proxy_percent': 0.0, 'fuel_tonnes_saved': 0.0, 'co2_tonnes_saved': 0.0, 'eta_hours_saved': 0.0, 'proof': {'mode': selected_mode, 'request_id': request_id}, 'mode_explanation': {'summary': 'Shortest distance path (A*). No weather data used.'}}}
            safe_payload = sanitize_for_json(distance_payload)
            ROUTE_RESPONSE_CACHE[route_cache_key] = {'created_at': time.time(), 'payload': safe_payload}
            _prune_cache_entries(ROUTE_RESPONSE_CACHE, ROUTE_CACHE_MAX_ENTRIES)
            logger.info('request_id=%s distance_mode_fast_return dist_km=%.1f', request_id, dist_km)
            await websocket.send_str(json.dumps(safe_payload))
            return
        await _progress(20, 'Sizing route corridor from weather...')
        corridor_radius_km, corridor_meta = await choose_corridor_radius_km(a_star_path)
        await _progress(30, 'Building navigation subgraph...')
        subgraph = build_subgraph(GLOBAL_GRAPH, GLOBAL_TREE, GLOBAL_NODE_ARRAY, a_star_path, radius_km=corridor_radius_km).to_directed()
        if not nx.has_path(subgraph, start_node, end_node):
            wider_radius = min(corridor_radius_km * 2, CORRIDOR_MAX_RADIUS_KM)
            logger.info('request_id=%s subgraph_disconnected initial_radius_km=%s widening_to_km=%s', request_id, corridor_radius_km, wider_radius)
            subgraph = build_subgraph(GLOBAL_GRAPH, GLOBAL_TREE, GLOBAL_NODE_ARRAY, a_star_path, radius_km=wider_radius).to_directed()
            if not nx.has_path(subgraph, start_node, end_node):
                logger.info('request_id=%s subgraph_still_disconnected fallback=full_graph_astar', request_id)
                try:
                    seed_path = nx.astar_path(GLOBAL_GRAPH, start_node, end_node, heuristic=_gc_heuristic_km, weight='distance')
                except nx.NetworkXNoPath:
                    seed_path = a_star_path
                subgraph = build_subgraph(GLOBAL_GRAPH, GLOBAL_TREE, GLOBAL_NODE_ARRAY, seed_path, radius_km=wider_radius).to_directed()
        logger.info('request_id=%s subgraph_nodes=%s subgraph_edges=%s connected=%s', request_id, subgraph.number_of_nodes(), subgraph.number_of_edges(), nx.has_path(subgraph, start_node, end_node))
        if nx.has_path(subgraph, start_node, end_node):
            try:
                subgraph_astar_path = nx.astar_path(subgraph, start_node, end_node, heuristic=_gc_heuristic_km, weight='distance')
                subgraph_astar_dist = calculate_total_nautical_distance(subgraph_astar_path) * 1.852
                if subgraph_astar_dist < astar_total_distance_km:
                    logger.info('request_id=%s subgraph_astar shorter: %.1f km → %.1f km', request_id, astar_total_distance_km, subgraph_astar_dist)
                    a_star_path = subgraph_astar_path
                    astar_total_distance_km = subgraph_astar_dist
            except nx.NetworkXNoPath:
                pass
        smooth_astarpath = constrained_smooth_path(a_star_path, points_per_edge=CONSTRAINED_SMOOTH_POINTS_PER_EDGE, max_points=CONSTRAINED_SMOOTH_MAX_POINTS)
        t0 = time.perf_counter()
        arrival_hours = estimate_node_arrival_hours(subgraph, end_node, astar_total_distance_km)
        logger.info('request_id=%s estimate_node_arrival_hours=%.2fs nodes=%s', request_id, time.perf_counter() - t0, subgraph.number_of_nodes())
        await _progress(40, 'Fetching weather & ocean data (GFS + CMEMS)…')
        t0 = time.perf_counter()
        forecast_hours = 72
        import asyncio as _asyncio
        async def _weather_heartbeat(stop_event):
            pct = 42
            steps = ['Downloading GFS wind & wave forecast…', 'Fetching CMEMS ocean current data…', 'Building weather corridor grid…', 'Interpolating weather to route nodes…', 'Analysing severe weather cells…']
            step_idx = 0
            while not stop_event.is_set() and pct < 74:
                await _asyncio.sleep(6)
                if stop_event.is_set():
                    break
                pct = min(pct + 4, 74)
                await _progress(pct, steps[step_idx % len(steps)])
                step_idx += 1
        _stop_evt = _asyncio.Event()
        _hb_task = _asyncio.create_task(_weather_heartbeat(_stop_evt))
        try:
            weather_context = await get_or_build_weather_context(subgraph, forecast_hours)
        finally:
            _stop_evt.set()
            await _hb_task
        threading.Thread(
            target=export_weather_to_excel, 
            args=(weather_context.get('grid_points', []), weather_context.get('current_weather_lookup', {}), weather_context.get('current_marine_lookup', {})),
            daemon=True
        ).start()
        logger.info('request_id=%s weather_context_build=%.2fs grid_points=%s (Excel export started in background)', request_id, time.perf_counter() - t0, len(weather_context.get('grid_points', [])))
        await _progress(75, 'Running route optimisation…')
        optimized_subgraph = subgraph.copy()
        edge_diagnostics = {}
        severity_points = weather_context.get('severity_points', [])
        if selected_mode == 'distance':
            logger.info('request_id=%s mode=distance using_turn_aware_subgraph_path', request_id)
            optimized_path = find_turn_aware_path(subgraph, start_node, end_node, edge_weight_key='distance', turn_penalty=TURN_PENALTY_DISTANCE_KM, max_turn_deg=MAX_ROUTE_TURN_DEG) or a_star_path
            smooth_path = constrained_smooth_path(optimized_path, points_per_edge=CONSTRAINED_SMOOTH_POINTS_PER_EDGE, max_points=CONSTRAINED_SMOOTH_MAX_POINTS)
            edge_diagnostics = {'total_edges': len(subgraph.edges()), 'edge_sample_points_mode': 'none_distance_mode_turn_aware', 'avg_edge_samples': 0.0, 'weather_cube_points': len(weather_context.get('grid_points', [])), 'high_weather_edges': 0, 'no_go_edges_blocked': 0, 'missing_weather_edges_blocked': 0}
        else:
            t0 = time.perf_counter()
            iter_result = optimize_path_with_iterative_refinement(subgraph=subgraph, start_node=start_node, end_node=end_node, objective_profile=objective_profile, selected_mode=selected_mode, total_distance_km=astar_total_distance_km, weather_context=weather_context, initial_arrival_hours=arrival_hours, initial_path=a_star_path)
            logger.info('request_id=%s iterative_optimization=%.2fs', request_id, time.perf_counter() - t0)
            optimized_subgraph = iter_result['graph']
            edge_diagnostics = iter_result['edge_diagnostics']
            severity_points = iter_result['severity_points']
            weather_context = iter_result['weather_context']
            optimized_path = iter_result['path'] or a_star_path
            if not optimized_path:
                try:
                    optimized_path = nx.dijkstra_path(subgraph, start_node, end_node, weight='distance')
                    logger.warning('request_id=%s fallback=distance_dijkstra', request_id)
                except nx.NetworkXNoPath:
                    optimized_path = a_star_path
                    logger.warning('request_id=%s fallback=baseline_geodesic', request_id)
            smooth_path = constrained_smooth_path(optimized_path, points_per_edge=CONSTRAINED_SMOOTH_POINTS_PER_EDGE, max_points=CONSTRAINED_SMOOTH_MAX_POINTS)
        await _progress(90, 'Computing route metrics & weather overlay…')
        new_smooth_path = [(node[1], node[0]) for node in smooth_path]
        new_astar = [(node[1], node[0]) for node in smooth_astarpath]
        marker_start_latlon = (float(start_coords[0]), float(start_coords[1]))
        marker_end_latlon = (float(end_coords[0]), float(end_coords[1]))
        new_smooth_path = attach_endpoint_connectors(new_smooth_path, marker_start_latlon, marker_end_latlon)
        new_astar = attach_endpoint_connectors(new_astar, marker_start_latlon, marker_end_latlon)
        current_weather_lookup = weather_context.get('current_weather_lookup', {})
        current_marine_lookup = weather_context.get('current_marine_lookup', {})
        grid_points = weather_context.get('grid_points', [])
        t0 = time.perf_counter()
        weather_info_list, astar_weather_info = await asyncio.gather(asyncio.to_thread(build_weather_info_from_context, new_smooth_path, weather_context), asyncio.to_thread(build_weather_info_from_context, new_astar, weather_context))
        logger.info('request_id=%s build_weather_info=%.2fs opt_points=%s astar_points=%s', request_id, time.perf_counter() - t0, len(weather_info_list), len(astar_weather_info))
        grid_meta = create_wind_grid(grid_points, current_weather_lookup, bbox_padding_deg=WIND_GRID_PADDING_DEG, grid_res=WIND_GRID_RES_DEG)
        if grid_meta[0] is not None:
            min_lon, max_lon, min_lat, max_lat, grid_lons, grid_lats, grid_u, grid_v = grid_meta
            wind_grid = {'min_lon': min_lon, 'max_lon': max_lon, 'min_lat': min_lat, 'max_lat': max_lat, 'lons': grid_lons, 'lats': grid_lats, 'u': grid_u, 'v': grid_v}
        else:
            wind_grid = None
        wind_field = build_vector_grid_from_lookups(grid_points, current_weather_lookup, current_marine_lookup, grid_rows=WIND_FIELD_GRID_ROWS, grid_cols=WIND_FIELD_GRID_COLS)
        if not severity_points:
            severity_points = build_severity_points_from_lookups(grid_points, current_weather_lookup, current_marine_lookup, max_points=800)
        severity_grid = build_severity_grid_payload(severity_points, rows=SEVERITY_GRID_ROWS, cols=SEVERITY_GRID_COLS)
        if severity_grid:
            logger.info('request_id=%s severity_grid rows=%s cols=%s', request_id, severity_grid.get('rows'), severity_grid.get('cols'))
        optimized_distance_km = calculate_total_nautical_distance(new_smooth_path) * 1.852
        astar_distance_km = calculate_total_nautical_distance(new_astar) * 1.852
        optimized_metrics = summarize_route_metrics('optimized', optimized_distance_km, weather_info_list)
        astar_metrics = summarize_route_metrics('astar', astar_distance_km, astar_weather_info)
        optimized_fuel_proxy = estimate_route_fuel_proxy(new_smooth_path, weather_info_list)
        astar_fuel_proxy = estimate_route_fuel_proxy(new_astar, astar_weather_info)
        fuel_saved = round(astar_fuel_proxy - optimized_fuel_proxy, 3)
        fuel_saved_percent = round(fuel_saved / astar_fuel_proxy * 100.0, 2) if astar_fuel_proxy else 0.0
        optimized_metrics['fuel_proxy'] = optimized_fuel_proxy
        astar_metrics['fuel_proxy'] = astar_fuel_proxy
        opt_eta_h = calculate_eta_hours(new_smooth_path, weather_info_list)
        astar_eta_h = calculate_eta_hours(new_astar, astar_weather_info)
        opt_polar = calculate_fuel_and_co2(new_smooth_path, weather_info_list)
        astar_polar = calculate_fuel_and_co2(new_astar, astar_weather_info)
        opt_fuel_t = opt_polar['fuel_tonnes']
        astar_fuel_t = astar_polar['fuel_tonnes']
        _co2_factor = 3.114
        opt_co2_t = round(opt_fuel_t * _co2_factor, 2)
        astar_co2_t = round(astar_fuel_t * _co2_factor, 2)
        optimized_metrics['eta_hours'] = opt_eta_h
        optimized_metrics['eta_days'] = round(opt_eta_h / 24.0, 2)
        optimized_metrics['fuel_tonnes'] = opt_fuel_t
        optimized_metrics['co2_tonnes'] = opt_co2_t
        astar_metrics['eta_hours'] = astar_eta_h
        astar_metrics['eta_days'] = round(astar_eta_h / 24.0, 2)
        astar_metrics['fuel_tonnes'] = astar_fuel_t
        astar_metrics['co2_tonnes'] = astar_co2_t
        fuel_tonnes_saved = round(astar_fuel_t - opt_fuel_t, 2)
        co2_tonnes_saved = round(astar_co2_t - opt_co2_t, 2)
        eta_hours_saved = round(astar_eta_h - opt_eta_h, 2)
        astar_violations = count_route_limit_violations(astar_weather_info)
        optimized_violations = count_route_limit_violations(weather_info_list)
        distance_saved = round(astar_metrics['distance_km'] - optimized_metrics['distance_km'], 3)
        percent_saved = round(distance_saved / astar_metrics['distance_km'] * 100.0, 2) if astar_metrics['distance_km'] else 0.0
        objective_graph = optimized_subgraph if selected_mode != 'distance' else subgraph
        astar_objective_score = compute_path_cost(objective_graph, a_star_path, weight_key='weight')
        optimized_objective_score = compute_path_cost(objective_graph, optimized_path, weight_key='weight')
        proof = {'mode': selected_mode, 'request_id': request_id, 'corridor_radius_km': corridor_radius_km, 'corridor_selection': corridor_meta, 'objective_score_astar': astar_objective_score, 'objective_score_optimized': optimized_objective_score, 'objective_improvement_pct': round((astar_objective_score - optimized_objective_score) / astar_objective_score * 100.0, 3) if astar_objective_score else 0.0, 'astar_violations': astar_violations, 'optimized_violations': optimized_violations, 'accepted_route': 'turn_aware_shortest_path', 'acceptance_reason': 'minimum_cost_with_turn_penalty_and_turn_cap'}
        if selected_mode == 'balanced' and BALANCED_ACCEPTANCE_ENABLED:
            risk_gain_pts = (_to_float_or_none(astar_metrics.get('risk_score')) or 0.0) - (_to_float_or_none(optimized_metrics.get('risk_score')) or 0.0)
            distance_increase_pct = (optimized_metrics['distance_km'] - astar_metrics['distance_km']) / astar_metrics['distance_km'] * 100.0 if astar_metrics['distance_km'] else 0.0
            eta_increase_hours = opt_eta_h - astar_eta_h
            fuel_increase_pct = (opt_fuel_t - astar_fuel_t) / astar_fuel_t * 100.0 if astar_fuel_t else 0.0
            commercial_regression = distance_increase_pct > BALANCED_MAX_DISTANCE_INCREASE_PCT or eta_increase_hours > BALANCED_MAX_ETA_INCREASE_HOURS or fuel_increase_pct > BALANCED_MAX_FUEL_INCREASE_PCT
            risk_gain_insufficient = risk_gain_pts < BALANCED_MIN_RISK_GAIN_PTS_FOR_REGRESSION
            proof['balanced_acceptance'] = {'distance_increase_pct': round(distance_increase_pct, 3), 'eta_increase_hours': round(eta_increase_hours, 3), 'fuel_increase_pct': round(fuel_increase_pct, 3), 'risk_gain_pts': round(risk_gain_pts, 3), 'thresholds': {'max_distance_increase_pct': BALANCED_MAX_DISTANCE_INCREASE_PCT, 'max_eta_increase_hours': BALANCED_MAX_ETA_INCREASE_HOURS, 'max_fuel_increase_pct': BALANCED_MAX_FUEL_INCREASE_PCT, 'min_risk_gain_pts': BALANCED_MIN_RISK_GAIN_PTS_FOR_REGRESSION}}
            if commercial_regression and risk_gain_insufficient:
                logger.info('request_id=%s balanced_guard fallback=astar dist+%.2f%% eta+%.2fh fuel+%.2f%% risk_gain=%.2fpts', request_id, distance_increase_pct, eta_increase_hours, fuel_increase_pct, risk_gain_pts)
                optimized_path = a_star_path
                smooth_path = smooth_astarpath
                weather_info_list = astar_weather_info
                new_smooth_path = new_astar
                optimized_distance_km = astar_distance_km
                optimized_metrics = dict(astar_metrics)
                optimized_metrics['label'] = 'optimized_fallback_astar'
                optimized_metrics['fuel_proxy'] = astar_fuel_proxy
                distance_saved = 0.0
                percent_saved = 0.0
                fuel_saved = 0.0
                fuel_saved_percent = 0.0
                fuel_tonnes_saved = 0.0
                co2_tonnes_saved = 0.0
                eta_hours_saved = 0.0
                optimized_violations = astar_violations
                optimized_objective_score = astar_objective_score
                proof['accepted_route'] = 'astar_fallback'
                proof['acceptance_reason'] = 'balanced_commercial_guard'
                proof['objective_score_optimized'] = optimized_objective_score
                proof['objective_improvement_pct'] = 0.0
            dist_increase_pct = (optimized_metrics['distance_km'] - astar_metrics['distance_km']) / astar_metrics['distance_km'] * 100.0 if astar_metrics['distance_km'] else 0.0
            risk_gain_pts = round(astar_metrics.get('risk_score', 0.0) - optimized_metrics.get('risk_score', 0.0), 2)
            safety_not_improved = risk_gain_pts <= 0.05
            marginal_safety_tradeoff = dist_increase_pct > 0.1 and risk_gain_pts < 1.0
            extreme_deviation = dist_increase_pct > 15.0 and risk_gain_pts < 5.0
            if safety_not_improved or marginal_safety_tradeoff or extreme_deviation:
                logger.info('request_id=%s safety_guard fallback=astar reason=%s', request_id, 'insufficient_gain' if extreme_deviation else 'no_improvement')
                optimized_path = a_star_path
                smooth_path = smooth_astarpath
                weather_info_list = astar_weather_info
                new_smooth_path = new_astar
                optimized_distance_km = astar_distance_km
                optimized_metrics = dict(astar_metrics)
                optimized_metrics['label'] = 'safety_fallback_astar'
                proof['accepted_route'] = 'astar_fallback'
                proof['acceptance_reason'] = 'safety_commercial_guard'
                optimized_metrics = dict(astar_metrics)
                optimized_metrics['label'] = 'optimized_fallback_astar'
                optimized_metrics['fuel_proxy'] = astar_fuel_proxy
                distance_saved = 0.0
                percent_saved = 0.0
                fuel_saved = 0.0
                fuel_saved_percent = 0.0
                fuel_tonnes_saved = 0.0
                co2_tonnes_saved = 0.0
                eta_hours_saved = 0.0
                optimized_violations = astar_violations
                optimized_objective_score = astar_objective_score
                proof['accepted_route'] = 'astar_fallback'
                proof['acceptance_reason'] = 'safety_not_improved'
                proof['objective_score_optimized'] = optimized_objective_score
                proof['objective_improvement_pct'] = 0.0
        pareto_routes = []
        if PARETO_ROUTE_ENABLED and selected_mode != 'distance':
            for label, profile in PARETO_WEIGHT_SETS.items():
                pareto_mode = 'safety' if label == 'safest' else 'balanced'
                pareto_result = optimize_path_with_iterative_refinement(subgraph=subgraph, start_node=start_node, end_node=end_node, objective_profile=profile, selected_mode=pareto_mode, total_distance_km=astar_total_distance_km, weather_context=weather_context, initial_arrival_hours=arrival_hours, initial_path=a_star_path)
                pareto_path_nodes = pareto_result['path'] or a_star_path
                pareto_path_latlon = [(node[1], node[0]) for node in pareto_path_nodes]
                pareto_weather = build_weather_info_from_context(pareto_path_latlon, weather_context)
                pareto_routes.append({'label': label, 'path': pareto_path_latlon, 'distance_km': round(calculate_total_nautical_distance(pareto_path_latlon) * 1.852, 3), 'objective_score': compute_path_cost(pareto_result['graph'], pareto_path_nodes, weight_key='weight'), 'metrics': summarize_route_metrics(label, calculate_total_nautical_distance(pareto_path_latlon) * 1.852, pareto_weather)})
        mode_explanation = build_mode_explanation(selected_mode, objective_profile, astar_metrics, optimized_metrics, fuel_saved, fuel_saved_percent, edge_diagnostics=edge_diagnostics)
        def _bg_llm_analysis(voyage_meta, pdf_path, loop_to_use):
            try:
                xl_path = os.path.join(os.path.dirname(__file__), 'route_weather_analysis.xlsx')
                # Wait briefly for Excel to finish if it's still writing
                for _ in range(20):
                    if os.path.exists(xl_path): break
                    time.sleep(0.5)
                
                if os.path.exists(xl_path):
                    voyage_analyzer.run_full_analysis(xl_path, voyage_meta, pdf_path)
                    msg = json.dumps({'type': 'report_ready', 'report_url': f"/reports/{os.path.basename(pdf_path)}"})
                    asyncio.run_coroutine_threadsafe(websocket.send_str(msg), loop_to_use)
            except Exception as exc:
                logger.warning('request_id=%s bg_analysis_failed error=%s', request_id, exc)

        report_pdf_path = os.path.join(os.path.dirname(__file__), os.getenv('AI_REPORT_NAME', 'voyage_report.pdf'))
        threading.Thread(
            target=_bg_llm_analysis, 
            args=(optimized_metrics, report_pdf_path, asyncio.get_event_loop()),
            daemon=True
        ).start()
        excel_path = None
        payload = {'type': 'final', 'request_id': request_id, 'path': new_smooth_path, 'weather': weather_info_list, 'optimized_weather': weather_info_list, 'astar_weather': astar_weather_info, 'severity': severity_points, 'severity_grid': severity_grid, 'wind_field': wind_field, 'wind_grid': wind_grid, 'distance': optimized_distance_km, 'apath': new_astar, 'mode': selected_mode, 'excel_export': excel_path, 'alternatives': pareto_routes, 'metrics': {'astar': astar_metrics, 'optimized': optimized_metrics, 'distance_saved_km': distance_saved, 'distance_saved_percent': percent_saved, 'fuel_saved_proxy': fuel_saved, 'fuel_saved_proxy_percent': fuel_saved_percent, 'fuel_tonnes_saved': fuel_tonnes_saved, 'co2_tonnes_saved': co2_tonnes_saved, 'eta_hours_saved': eta_hours_saved, 'proof': proof, 'mode_explanation': mode_explanation}}
        t0 = time.perf_counter()
        safe_payload = sanitize_for_json(payload)
        logger.info('request_id=%s sanitize_for_json=%.2fs', request_id, time.perf_counter() - t0)
        ROUTE_RESPONSE_CACHE[route_cache_key] = {'created_at': time.time(), 'payload': safe_payload}
        _prune_cache_entries(ROUTE_RESPONSE_CACHE, ROUTE_CACHE_MAX_ENTRIES)
        logger.info('request_id=%s sending_response path_points=%s astar_points=%s weather_points=%s alternatives=%s', request_id, len(new_smooth_path), len(new_astar), len(weather_info_list), len(pareto_routes))
        await websocket.send_str(json.dumps(safe_payload, allow_nan=False))
    except Exception as e:
        logger.exception('request_id=%s handle_navigation error=%s', request_id, e)
        error_payload = sanitize_for_json({'type': 'error', 'message': str(e)})
        await websocket.send_str(json.dumps(error_payload, allow_nan=False))
        raise
def export_weather_to_excel(grid_points, current_weather_lookup, current_marine_lookup, output_path=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning('[excel] openpyxl not installed; install with: pip install openpyxl')
        return None
    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), 'route_weather_analysis.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weather Analysis'
    headers = ['Latitude', 'Longitude', 'Wind Speed (km/h)', 'Wind Direction (°)', 'Precipitation (mm/h)', 'Visibility (m)', 'Wave Height (m)', 'Wave Direction (°)', 'Current Velocity (m/s)', 'Current Direction (°)', 'Wave Sev ×0.40', 'Wind Sev ×0.30', 'Vis Risk ×0.20', 'Precip Sev ×0.10', 'Severity Score (0-100)', 'Severity Calculation']
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    def _risk_fill(score):
        s = max(0.0, min(score / 100.0, 1.0))
        if s < 0.33:
            r, g = (int(s / 0.33 * 255), 200)
        elif s < 0.66:
            r, g = (255, int((1 - (s - 0.33) / 0.33) * 200))
        else:
            r, g = (220, 0)
        return PatternFill(start_color=f'{r:02X}{g:02X}00', end_color=f'{r:02X}{g:02X}00', fill_type='solid')
    for row_idx, loc in enumerate(grid_points, 2):
        lat, lon = (float(loc[0]), float(loc[1]))
        cw = current_weather_lookup.get(loc, {}).get('current', {})
        cm = current_marine_lookup.get(loc, {}).get('current', {})
        wind = float(cw.get('wind_speed_10m', 0) or 0)
        wdir = float(cw.get('wind_direction_10m', 0) or 0)
        prcp = float(cw.get('precipitation', 0) or 0)
        vis = float(cw.get('visibility', 10000) or 10000)
        wave = float(cm.get('wave_height', 0) or 0)
        waved = float(cm.get('wave_direction', 0) or 0)
        cvel = float(cm.get('ocean_current_velocity', 0) or 0)
        cdir = float(cm.get('ocean_current_direction', 0) or 0)
        wave_sev = min(wave / 4.0, 1.0)
        wind_sev = min(wind / 35.0, 1.0)
        precip_sev = min(prcp / 10.0, 1.0)
        vis_risk = max(0.0, 1.0 - min(vis / 10000.0, 1.0))
        w_wave = round(0.4 * wave_sev, 4)
        w_wind = round(0.3 * wind_sev, 4)
        w_vis = round(0.2 * vis_risk, 4)
        w_prcp = round(0.1 * precip_sev, 4)
        score = round((w_wave + w_wind + w_vis + w_prcp) * 100.0, 2)
        formula = f'(0.40 × {wave_sev:.3f}) + (0.30 × {wind_sev:.3f}) + (0.20 × {vis_risk:.3f}) + (0.10 × {precip_sev:.3f}) = {score / 100:.4f} → {score}'
        row = [lat, lon, round(wind, 2), round(wdir, 1), round(prcp, 3), round(vis, 0), round(wave, 3), round(waved, 1), round(cvel, 3), round(cdir, 1), w_wave, w_wind, w_vis, w_prcp, score, formula]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='center')
        score_cell = ws.cell(row=row_idx, column=15)
        score_cell.fill = _risk_fill(score)
        score_cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('Formula Reference')
    formula_rows = [['Factor', 'Formula', 'Max Risk Threshold', 'Weight'], ['Wave Height', 'min(wave_m / 4.0,  1.0)', '4.0 m', '40%'], ['Wind Speed', 'min(wind_kmh / 35.0, 1.0)', '35 km/h (Beaufort 6)', '30%'], ['Visibility', '1 - min(vis_m / 10000, 1)', '0 m (zero vis)', '20%'], ['Precipitation', 'min(precip_mm / 10.0, 1.0)', '10 mm/h (heavy rain)', '10%'], [], ['Severity Score', '= (wave×0.40 + wind×0.30 + vis×0.20 + precip×0.10) × 100'], ['Range', '0 = completely safe,  100 = maximum danger']]
    for fr in formula_rows:
        ws2.append(fr)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
    wb.save(output_path)
    logger.info('[excel] Weather analysis saved path=%s rows=%s', output_path, len(grid_points))
    return output_path
def init_globals():
    logger.info('WebSocket server is starting on ws://localhost:5000')
    global GLOBAL_GRAPH, GLOBAL_TREE, GLOBAL_NODE_ARRAY
    logger.info('Open-Meteo limiter config: %s/min %s/hour %s/day batch=%s retries=%s', OPEN_METEO_MAX_CALLS_PER_MIN, OPEN_METEO_MAX_CALLS_PER_HOUR, OPEN_METEO_MAX_CALLS_PER_DAY, OPEN_METEO_BATCH_SIZE, OPEN_METEO_MAX_RETRIES)
    logger.info('Weather cube config: model_run=%sh spacing=%sdeg padding=%sdeg max_points=%s interp_neighbors=%s adaptive_edge_sampling=%s', WEATHER_MODEL_RUN_INTERVAL_HOURS, CORRIDOR_GRID_SPACING_DEG, CORRIDOR_GRID_PADDING_DEG, CORRIDOR_GRID_MAX_POINTS, CORRIDOR_INTERP_NEIGHBORS, 'on' if ADAPTIVE_EDGE_SAMPLING_ENABLED else 'off')
    logger.info('Cache config: route_cache_max=%s weather_cache_max=%s', ROUTE_CACHE_MAX_ENTRIES, WEATHER_CACHE_MAX_ENTRIES)
    logger.info('DEBUG_MODE=%s', 'on' if DEBUG_MODE else 'off')
    startup_t0 = time.perf_counter()
    GLOBAL_GRAPH = load_navigation_graph(GRAPH_PATH)
    GLOBAL_TREE, GLOBAL_NODE_ARRAY = build_spatial_index(GLOBAL_GRAPH)
    logger.info('Graph preload complete in %.2fs', time.perf_counter() - startup_t0)